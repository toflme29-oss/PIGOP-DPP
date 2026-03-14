"""Endpoints — Control de Oficios Recibidos."""

import io
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.dependencies import get_current_active_user
from app.core.database import get_db
from app.core.exceptions import ConflictError, NotFoundError
from app.crud.oficio import crud_oficio
from app.models.oficio import OficioRecibido
from app.models.user import Usuario
from app.schemas.common import MessageResponse, PaginatedResponse
from app.schemas.oficio import OficioCreate, OficioResponse, OficioUpdate

router = APIRouter()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _build_filters(
    current_user: Usuario,
    fecha_desde: Optional[date],
    fecha_hasta: Optional[date],
    dependencia: Optional[str],
    busqueda: Optional[str],
):
    filters = []
    # Multi-tenancy: superadmin ve todos, otros solo su cliente
    if current_user.rol != "superadmin":
        filters.append(OficioRecibido.cliente_id == str(current_user.cliente_id))

    if fecha_desde:
        filters.append(OficioRecibido.fecha_oficio >= fecha_desde)
    if fecha_hasta:
        filters.append(OficioRecibido.fecha_oficio <= fecha_hasta)
    if dependencia:
        filters.append(OficioRecibido.dependencia.ilike(f"%{dependencia}%"))
    if busqueda:
        pattern = f"%{busqueda}%"
        filters.append(
            OficioRecibido.numero_oficio.ilike(pattern)
            | OficioRecibido.remitente.ilike(pattern)
            | OficioRecibido.asunto.ilike(pattern)
        )
    return filters


def _to_response(oficio: OficioRecibido) -> OficioResponse:
    data = OficioResponse.model_validate(oficio)
    if oficio.registrador:
        data.registrador_nombre = oficio.registrador.nombre_completo
    return data


# ── CRUD ──────────────────────────────────────────────────────────────────────

@router.post("/", response_model=OficioResponse, status_code=201, summary="Registrar oficio recibido")
async def crear_oficio(
    data: OficioCreate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    cliente_id = str(current_user.cliente_id) if current_user.rol != "superadmin" else str(current_user.cliente_id)

    # Validar duplicado
    existente = await crud_oficio.get_by_numero_oficio(db, data.numero_oficio, cliente_id)
    if existente:
        raise ConflictError(f"Ya existe un oficio con número '{data.numero_oficio}'.")

    folio = await crud_oficio.get_next_folio(db, cliente_id)

    obj_data = data.model_dump()
    obj_data["folio"] = folio
    obj_data["cliente_id"] = cliente_id
    obj_data["registrado_por"] = str(current_user.id)

    oficio = await crud_oficio.create(db, obj_in=obj_data)
    return _to_response(oficio)


@router.get("/", response_model=PaginatedResponse[OficioResponse], summary="Listar oficios recibidos")
async def listar_oficios(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    fecha_desde: Optional[date] = Query(None, description="Fecha inicio (YYYY-MM-DD)"),
    fecha_hasta: Optional[date] = Query(None, description="Fecha fin (YYYY-MM-DD)"),
    dependencia: Optional[str] = Query(None),
    busqueda: Optional[str] = Query(None, description="Búsqueda en número, remitente o asunto"),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    filters = _build_filters(current_user, fecha_desde, fecha_hasta, dependencia, busqueda)
    items = await crud_oficio.get_multi_ordered(db, skip=skip, limit=limit, filters=filters)
    total = await crud_oficio.count(db, filters=filters)
    return PaginatedResponse(
        items=[_to_response(o) for o in items],
        total=total,
        skip=skip,
        limit=limit,
    )


@router.get("/export", summary="Exportar oficios a Excel")
async def exportar_oficios(
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    dependencia: Optional[str] = Query(None),
    busqueda: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Genera y descarga un archivo Excel con los oficios filtrados."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    filters = _build_filters(current_user, fecha_desde, fecha_hasta, dependencia, busqueda)
    oficios = await crud_oficio.get_multi_ordered(db, skip=0, limit=10000, filters=filters)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Oficios Recibidos"

    # ── Estilos ───────────────────────────────────────────────────────────────
    guinda_fill = PatternFill("solid", fgColor="911A3A")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="top", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    alt_fill = PatternFill("solid", fgColor="FDF2F4")

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "Control de Oficios Recibidos — PIGOP"
    title_cell.font = Font(bold=True, size=14, color="911A3A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtítulo con rango de fechas
    ws.merge_cells("A2:I2")
    rango = "Todos los registros"
    if fecha_desde and fecha_hasta:
        rango = f"Del {fecha_desde.isoformat()} al {fecha_hasta.isoformat()}"
    elif fecha_desde:
        rango = f"Desde {fecha_desde.isoformat()}"
    elif fecha_hasta:
        rango = f"Hasta {fecha_hasta.isoformat()}"
    ws["A2"].value = rango
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Encabezados ───────────────────────────────────────────────────────────
    columns = [
        ("Folio", 8),
        ("No. Oficio", 20),
        ("Remitente", 30),
        ("Dependencia", 30),
        ("Asunto", 40),
        ("Descripción", 35),
        ("Fecha Oficio", 14),
        ("Fecha Registro", 18),
        ("Observaciones", 30),
    ]

    header_row = 4
    for col_idx, (col_name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = guinda_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = width

    # ── Datos ─────────────────────────────────────────────────────────────────
    for row_idx, oficio in enumerate(oficios, start=header_row + 1):
        values = [
            oficio.folio,
            oficio.numero_oficio,
            oficio.remitente,
            oficio.dependencia,
            oficio.asunto,
            oficio.descripcion or "",
            oficio.fecha_oficio.isoformat() if oficio.fecha_oficio else "",
            oficio.fecha_registro.strftime("%Y-%m-%d %H:%M") if oficio.fecha_registro else "",
            oficio.observaciones or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin
            cell.alignment = wrap
            if (row_idx - header_row) % 2 == 0:
                cell.fill = alt_fill

    # ── Footer ────────────────────────────────────────────────────────────────
    footer_row = header_row + len(oficios) + 2
    ws.merge_cells(f"A{footer_row}:I{footer_row}")
    ws[f"A{footer_row}"].value = f"Total de oficios: {len(oficios)}"
    ws[f"A{footer_row}"].font = Font(bold=True, size=10, color="911A3A")

    # Serializar
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = "PIGOP_Oficios_Recibidos"
    if fecha_desde:
        filename += f"_{fecha_desde.isoformat()}"
    if fecha_hasta:
        filename += f"_{fecha_hasta.isoformat()}"
    filename += ".xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{oficio_id}", response_model=OficioResponse, summary="Detalle de oficio")
async def obtener_oficio(
    oficio_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    oficio = await crud_oficio.get(db, oficio_id)
    if not oficio:
        raise NotFoundError("Oficio no encontrado.")
    return _to_response(oficio)


@router.put("/{oficio_id}", response_model=OficioResponse, summary="Actualizar oficio")
async def actualizar_oficio(
    oficio_id: str,
    data: OficioUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    oficio = await crud_oficio.get(db, oficio_id)
    if not oficio:
        raise NotFoundError("Oficio no encontrado.")

    # Si cambia el número de oficio, validar que no haya duplicado
    if data.numero_oficio and data.numero_oficio != oficio.numero_oficio:
        existente = await crud_oficio.get_by_numero_oficio(
            db, data.numero_oficio, str(oficio.cliente_id)
        )
        if existente:
            raise ConflictError(f"Ya existe un oficio con número '{data.numero_oficio}'.")

    update_data = data.model_dump(exclude_unset=True)
    updated = await crud_oficio.update(db, db_obj=oficio, obj_in=update_data)
    return _to_response(updated)


@router.delete("/{oficio_id}", response_model=MessageResponse, summary="Eliminar oficio")
async def eliminar_oficio(
    oficio_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    oficio = await crud_oficio.delete(db, id=oficio_id)
    if not oficio:
        raise NotFoundError("Oficio no encontrado.")
    return MessageResponse(message="Oficio eliminado correctamente.")
