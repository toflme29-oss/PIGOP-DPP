"""
Endpoints para ingesta SAP — Importación de DEPPs desde archivo Excel/CSV.

Flujo de 3 pasos desde el frontend:
  1. POST /sap/import/preview  → sube archivo, retorna preview (sin crear nada)
  2. GET  /sap/import/template → descarga plantilla Excel con las columnas esperadas
  3. POST /sap/import/confirm  → con el log_id del preview, confirma y crea los DEPPs

Otros:
  GET  /sap/import/logs       → historial de importaciones
  GET  /sap/status            → estado de conexión SAP (mock/rfc/odata)
"""
import io
import uuid
from datetime import datetime, timezone
from typing import List, Optional

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.dependencies import get_current_active_user, get_current_admin
from app.core.database import get_db
from app.core.exceptions import BusinessError, NotFoundError
from app.models.sap import SAPImportLog
from app.models.user import Usuario
from app.services.sap_import_service import importar_depps_desde_archivo

router = APIRouter()


# ── Status SAP ────────────────────────────────────────────────────────────────

@router.get("/status", summary="Estado de la conexión SAP")
async def sap_status():
    """
    Retorna el modo de conexión SAP actual.
    Mientras SAP BASIS no habilite el acceso directo, el modo es 'archivo'.
    """
    return {
        "modo": "archivo",
        "disponible": True,
        "descripcion": "Importación via archivo Excel/CSV exportado de SAP GRP",
        "modos_disponibles": [
            {
                "modo": "archivo",
                "activo": True,
                "descripcion": "Excel/CSV exportado manualmente de SAP (FBL1N, ZDEPP, etc.)",
            },
            {
                "modo": "rfc",
                "activo": False,
                "descripcion": "Llamadas directas RFC/BAPI — requiere SAP NW RFC SDK y acceso de red",
            },
            {
                "modo": "odata",
                "activo": False,
                "descripcion": "REST via SAP Gateway — requiere SAP Gateway habilitado en el servidor",
            },
        ],
    }


# ── Template de importación ───────────────────────────────────────────────────

@router.get("/import/template", summary="Descargar plantilla Excel para importación SAP")
async def descargar_template(
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Descarga una plantilla Excel con las columnas esperadas para importar DEPPs.
    El analista llena esta plantilla con datos exportados de SAP y la sube a PIGOP.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DEPPs SAP"

        # Encabezados
        columns = [
            ("FOLIO_DEPP", "Folio único del DEPP (ej: DPP-2026-001)", True),
            ("UPP", "Código de Unidad Programática (ej: 007)", True),
            ("EJERCICIO", "Año fiscal (ej: 2026)", True),
            ("MES", "Mes del trámite (1-12)", False),
            ("CAPITULO", "Capítulo presupuestal (ej: 3000)", False),
            ("CONCEPTO", "Concepto presupuestal", False),
            ("PARTIDA", "Partida presupuestal (ej: 3231)", False),
            ("MONTO_TOTAL", "Importe total del DEPP en pesos MXN", True),
            ("BENEFICIARIO", "Nombre del proveedor o beneficiario", False),
            ("RFC_BENEFICIARIO", "RFC del proveedor o beneficiario", False),
            ("CLAVE_PRESUPUESTARIA", "Clave programática completa", False),
            ("FUENTE_FINANCIAMIENTO", "Fuente de financiamiento", False),
            ("TIPO_PAGO", "Modalidad de pago (TRANSFERENCIA, CHEQUE, etc.)", False),
            ("NRO_DOC_SAP", "Número de documento FI en SAP (referencia interna)", False),
            ("CLASIFICADOR", "Tipo de trámite (I.1, II.1, II.2, II.3, II.4)", False),
        ]

        # Estilo encabezado
        header_fill_req  = PatternFill("solid", fgColor="911A3A")
        header_fill_opt  = PatternFill("solid", fgColor="C0392B")
        header_font_req  = Font(bold=True, color="FFFFFF", size=10)
        header_font_opt  = Font(bold=False, color="FFFFFF", size=10)
        center           = Alignment(horizontal="center", vertical="center")
        thin             = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
        example_fill = PatternFill("solid", fgColor="FDF2F4")

        for col_idx, (col_name, desc, required) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill   = header_fill_req if required else header_fill_opt
            cell.font   = header_font_req if required else header_font_opt
            cell.alignment = center
            cell.border = thin
            ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 4, 18)

        # Fila de descripción (fila 2, en gris)
        desc_fill = PatternFill("solid", fgColor="F5F5F5")
        for col_idx, (col_name, desc, required) in enumerate(columns, start=1):
            cell = ws.cell(row=2, column=col_idx, value=f"{'*Requerido' if required else 'Opcional'}: {desc}")
            cell.fill = desc_fill
            cell.font = Font(italic=True, size=8, color="666666")
            cell.alignment = Alignment(wrap_text=True)
            cell.border = thin
        ws.row_dimensions[2].height = 30

        # Filas de ejemplo
        examples = [
            ["DPP-2026-001", "007", 2026, 2, 3000, 3200, 3231, 45800.00,
             "EMPRESA EJEMPLO SA DE CV", "EEJ200101ABC", "07-3-3000-3231",
             "Federal", "TRANSFERENCIA", "1000001234", "I.1"],
            ["DPP-2026-002", "015", 2026, 2, 2000, 2100, 2111, 12500.50,
             "PAPELERÍA DEL ESTADO", "PDE190505XYZ", "15-2-2000-2111",
             "Estatal", "TRANSFERENCIA", "1000001235", "II.4"],
        ]
        for row_idx, example_row in enumerate(examples, start=3):
            for col_idx, val in enumerate(example_row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill   = example_fill
                cell.font   = Font(size=9)
                cell.border = thin

        # Hoja de instrucciones
        ws_inst = wb.create_sheet("Instrucciones")
        instrucciones = [
            ("PIGOP — Plantilla de Importación SAP", True, 14),
            ("", False, 11),
            ("INSTRUCCIONES DE USO:", True, 11),
            ("1. Exporta los DEPPs de SAP usando la transacción correspondiente (FBL1N, ZDEPP u otra).", False, 10),
            ("2. Copia los datos en la hoja 'DEPPs SAP', respetando los encabezados.", False, 10),
            ("3. Las columnas marcadas con * son obligatorias (FOLIO_DEPP, UPP, EJERCICIO, MONTO_TOTAL).", False, 10),
            ("4. El sistema detecta automáticamente columnas con nombres alternativos.", False, 10),
            ("5. Si un DEPP con el mismo folio y ejercicio ya existe, se omite (no se duplica).", False, 10),
            ("6. Sube el archivo en PIGOP → Importación SAP → Cargar archivo.", False, 10),
            ("", False, 10),
            ("COLUMNAS ACEPTADAS (PIGOP acepta variantes):", True, 11),
            ("FOLIO_DEPP: folio, folio_depp, num_depp, numero_depp, documento", False, 10),
            ("UPP: upp, unidad_programatica, unidad_presupuestal, codigo_upp", False, 10),
            ("MONTO_TOTAL: monto_total, monto, importe, total, importe_total", False, 10),
            ("BENEFICIARIO: beneficiario, proveedor, razon_social, nombre_proveedor", False, 10),
            ("", False, 10),
            ("CLASIFICADORES VÁLIDOS:", True, 11),
            ("I.1  — Pago a proveedor con contrato (CFDI + CTT + MCL)", False, 10),
            ("II.1 — Reasignación de recursos (AUR)", False, 10),
            ("II.2 — Comisión oficial (FUC)", False, 10),
            ("II.3 — Transferencia directa (PCH)", False, 10),
            ("II.4 — Pago sin contrato (CFDI + MCL)", False, 10),
            ("", False, 10),
            ("Soporte: Dirección de Programación y Presupuesto — SFA Michoacán", False, 9),
        ]
        for row_idx, (text, bold, size) in enumerate(instrucciones, start=1):
            cell = ws_inst.cell(row=row_idx, column=1, value=text)
            cell.font = Font(bold=bold, size=size)
        ws_inst.column_dimensions["A"].width = 90

        # Serializar
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=PIGOP_Plantilla_SAP_2026.xlsx"
            },
        )

    except ImportError:
        # Fallback CSV si openpyxl no está instalado
        csv_content = (
            "FOLIO_DEPP,UPP,EJERCICIO,MES,CAPITULO,CONCEPTO,PARTIDA,"
            "MONTO_TOTAL,BENEFICIARIO,RFC_BENEFICIARIO,CLAVE_PRESUPUESTARIA,"
            "FUENTE_FINANCIAMIENTO,TIPO_PAGO,NRO_DOC_SAP,CLASIFICADOR\n"
            "DPP-2026-001,007,2026,2,3000,3200,3231,45800.00,"
            "EMPRESA EJEMPLO SA DE CV,EEJ200101ABC,07-3-3000-3231,"
            "Federal,TRANSFERENCIA,1000001234,I.1\n"
        )
        return StreamingResponse(
            io.BytesIO(csv_content.encode()),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=PIGOP_Plantilla_SAP_2026.csv"},
        )


# ── Preview (paso 1) ──────────────────────────────────────────────────────────

@router.post("/import/preview", summary="Paso 1: Previsualizar archivo antes de importar")
async def preview_importacion(
    file: UploadFile = File(..., description="Archivo Excel (.xlsx) o CSV exportado de SAP"),
    ejercicio: int = Query(2026, ge=2020, le=2050),
    mes: Optional[int] = Query(None, ge=1, le=12),
    upp_filtro: Optional[str] = Query(None, description="Filtrar solo la UPP indicada"),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Paso 1 del wizard de importación.
    Sube el archivo y retorna un preview de las primeras filas + mapeo de columnas detectado.
    NO crea ningún DEPP todavía.
    """
    if not file.filename:
        raise BusinessError("El archivo no tiene nombre.")
    ext = file.filename.lower().rsplit(".", 1)[-1]
    if ext not in ("xlsx", "xls", "csv"):
        raise BusinessError("Solo se aceptan archivos .xlsx, .xls o .csv exportados de SAP.")

    file_bytes = await file.read()
    if len(file_bytes) > 10 * 1024 * 1024:  # 10 MB máximo
        raise BusinessError("El archivo supera los 10 MB. Divide la exportación en rangos menores.")

    # Determinar cliente_id
    if current_user.rol == "superadmin":
        cliente_id = str(current_user.cliente_id) if current_user.cliente_id else "00000000-0000-0000-0000-000000000000"
    else:
        cliente_id = str(current_user.cliente_id)

    log = await importar_depps_desde_archivo(
        db=db,
        file_bytes=file_bytes,
        filename=file.filename,
        cliente_id=cliente_id,
        usuario_id=str(current_user.id),
        ejercicio=ejercicio,
        mes=mes,
        upp_filtro=upp_filtro,
        dry_run=True,   # Solo preview, no crea DEPPs
    )
    await db.commit()

    return {
        "log_id": log.id,
        "estado": log.estado,
        "nombre_archivo": log.nombre_archivo,
        "total_filas": log.total_filas,
        "preview": log.preview_data,
        "errores": log.errores_detalle,
    }


# ── Confirmar importación (paso 2) ────────────────────────────────────────────

@router.post("/import/confirmar/{log_id}", summary="Paso 2: Confirmar y ejecutar importación")
async def confirmar_importacion(
    log_id: str,
    file: UploadFile = File(..., description="Mismo archivo del paso de preview"),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Paso 2 del wizard. El usuario revisó el preview y confirma la importación.
    Crea los DEPPs en estado 'en_tramite'. Folios duplicados se omiten.
    """
    # Recuperar el log del preview para tomar los parámetros
    log_preview = await db.get(SAPImportLog, log_id)
    if not log_preview:
        raise NotFoundError("Log de importación")
    if log_preview.estado not in ("pendiente", "error_parcial"):
        raise BusinessError(
            f"Este log ya fue procesado (estado: {log_preview.estado}). "
            "Genera un nuevo preview para reimportar."
        )

    file_bytes = await file.read()

    if current_user.rol == "superadmin":
        cliente_id = str(log_preview.cliente_id)
    else:
        cliente_id = str(current_user.cliente_id)

    log = await importar_depps_desde_archivo(
        db=db,
        file_bytes=file_bytes,
        filename=log_preview.nombre_archivo or file.filename or "import.xlsx",
        cliente_id=cliente_id,
        usuario_id=str(current_user.id),
        ejercicio=log_preview.ejercicio,
        mes=log_preview.mes,
        upp_filtro=log_preview.upp_filtro,
        dry_run=False,   # ¡Crear DEPPs!
    )
    await db.commit()

    return {
        "log_id": log.id,
        "estado": log.estado,
        "total_filas": log.total_filas,
        "depps_creados": log.depps_creados,
        "depps_omitidos": log.depps_omitidos,
        "depps_error": log.depps_error,
        "errores": log.errores_detalle,
        "mensaje": (
            f"Importación completada: {log.depps_creados} DEPPs creados, "
            f"{log.depps_omitidos} omitidos (ya existían), "
            f"{log.depps_error} con error."
        ),
    }


# ── Importación directa (sin wizard) ─────────────────────────────────────────

@router.post("/import/archivo", summary="Importar DEPPs desde archivo (directo, sin preview)")
async def importar_archivo_directo(
    file: UploadFile = File(...),
    ejercicio: int = Query(2026, ge=2020, le=2050),
    mes: Optional[int] = Query(None, ge=1, le=12),
    upp_filtro: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Importación directa sin paso de preview. Para uso desde scripts o API externa."""
    if not file.filename:
        raise BusinessError("Archivo sin nombre.")
    file_bytes = await file.read()

    if current_user.rol == "superadmin":
        cliente_id = str(current_user.cliente_id) if current_user.cliente_id else "00000000-0000-0000-0000-000000000000"
    else:
        cliente_id = str(current_user.cliente_id)

    log = await importar_depps_desde_archivo(
        db=db,
        file_bytes=file_bytes,
        filename=file.filename,
        cliente_id=cliente_id,
        usuario_id=str(current_user.id),
        ejercicio=ejercicio,
        mes=mes,
        upp_filtro=upp_filtro,
        dry_run=False,
    )
    await db.commit()
    return log


# ── Historial de importaciones ────────────────────────────────────────────────

@router.get("/import/logs", summary="Historial de importaciones SAP")
async def listar_logs(
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Lista el historial de importaciones del cliente actual."""
    from sqlalchemy import select
    q = select(SAPImportLog).order_by(SAPImportLog.iniciado_en.desc()).limit(limit)

    if current_user.rol != "superadmin" and current_user.cliente_id:
        q = q.where(SAPImportLog.cliente_id == str(current_user.cliente_id))

    result = await db.execute(q)
    logs = result.scalars().all()

    return [
        {
            "id": log.id,
            "modo": log.modo,
            "nombre_archivo": log.nombre_archivo,
            "ejercicio": log.ejercicio,
            "mes": log.mes,
            "estado": log.estado,
            "total_filas": log.total_filas,
            "depps_creados": log.depps_creados,
            "depps_omitidos": log.depps_omitidos,
            "depps_error": log.depps_error,
            "iniciado_en": log.iniciado_en.isoformat() if log.iniciado_en else None,
            "completado_en": log.completado_en.isoformat() if log.completado_en else None,
        }
        for log in logs
    ]
