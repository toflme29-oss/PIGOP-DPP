import logging
import os
import uuid
from typing import List, Optional

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Body, Depends, File, Form, Query, Request, Response, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.dependencies import get_current_active_user
from app.core.database import get_db
from app.core.exceptions import BusinessError, ForbiddenError, NotFoundError
from app.crud.documento import crud_documento
from app.models.user import Usuario
from app.schemas.common import MessageResponse
from fastapi.responses import StreamingResponse, FileResponse
from app.schemas.documento import (
    ConfirmarTurnoInput,
    DevolucionInput,
    DevolucionResponse,
    DocumentoEmitidoCreate,
    DocumentoListResponse,
    DocumentoRecibidoCreate,
    DocumentoResponse,
    DocumentoUpdate,
    FirmaResponse,
    HistorialItemResponse,
    OficioEstructuradoResponse,
    PreviewOCRResponse,
    ProcesarOCRResponse,
    ReenvioInput,
    ESTADOS_EMITIDO,
    ESTADOS_RECIBIDO,
)
from app.services.correspondencia_service import (
    AREAS_DPP,
    PREFIJOS_FOLIO,
    PLANTILLAS_OFICIO,
    correspondencia_service,
    clasificar_oficio,
    detectar_plantilla,
    obtener_copias,
)

router = APIRouter()

UPLOAD_DIR = os.path.join(
    os.path.dirname(__file__), "..", "..", "..", "..", "uploads", "documentos"
)
os.makedirs(UPLOAD_DIR, exist_ok=True)

MIME_PERMITIDOS = {
    "application/pdf",
    "image/jpeg", "image/jpg", "image/png", "image/tiff", "image/webp",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "text/csv",
    "text/plain",
}
MAX_SIZE_MB = 20


def _assert_acceso(current_user: Usuario, cliente_id: str) -> None:
    if current_user.rol == "superadmin":
        return
    if str(current_user.cliente_id) != str(cliente_id):
        raise ForbiddenError("No tienes acceso a recursos de este cliente.")


# Jerarquía de áreas DPP — cada área ve sus propias + departamentos subordinados
AREA_JERARQUIA = {
    "DIR":  None,   # Director y Secretaría ven TODO (no se filtra)
    "SEC":  None,   # Bandeja de Secretaría del Director — también ve todo (rol=secretaria)
    "SCG":  ["SCG", "DREP", "DCP"],     # Subdir. Control y Gasto → sus departamentos
    "SPF":  ["SPF", "DASP", "DFNP"],    # Subdir. Presupuesto y Finanzas → sus departamentos
    "DREP": ["DREP"],                     # Depto. solo se ve a sí mismo
    "DCP":  ["DCP"],
    "DASP": ["DASP"],
    "DFNP": ["DFNP"],
}

# Áreas cuya asignación está restringida por rol. Las áreas aquí listadas solo
# pueden ser turnadas por los roles permitidos. El resto recibirá un 403.
AREAS_ASIGNACION_RESTRINGIDA = {
    "SEC": {
        "roles_permitidos": {"admin_cliente", "superadmin"},
        "mensaje": (
            "Solo el Director puede instruir a la Secretaría a contestar un oficio. "
            "Turna a tu subdirección o departamento responsable."
        ),
    },
}

# Roles con visibilidad total (ven todos los oficios)
_ROLES_VISION_TOTAL = {"superadmin", "admin_cliente", "secretaria", "asesor", "auditor", "consulta"}
# Roles con visibilidad restringida por area_codigo
_ROLES_VISION_AREA = {"subdirector", "jefe_depto", "analista"}


def _areas_visibles(user: Usuario) -> Optional[list[str]]:
    """
    Retorna la lista de area_turno que el usuario puede ver:
      - None: ve todo (Director, Secretaría, Superadmin, Asesor, Auditor, Consulta)
      - [códigos]: solo ve documentos con area_turno en esa lista
      - []: no ve nada (caso de subdirector/jefe sin area_codigo asignado)

    Reglas:
      - Subdirector: ve su área + departamentos subordinados (según AREA_JERARQUIA)
      - Jefe de Departamento / Analista: solo ve su propia área
      - Sin area_codigo asignado pero con rol restringido → NO VE NADA
        (se fuerza al admin a configurar el area)
    """
    if user.rol in _ROLES_VISION_TOTAL:
        return None

    if user.rol in _ROLES_VISION_AREA:
        area = getattr(user, 'area_codigo', None)
        if not area:
            # Rol restringido SIN area_codigo → NO VE NADA (en vez de ver todo).
            # Esto cierra un hueco: antes retornaba None (ver todo) por compat.
            logger.warning(
                f"Usuario {user.email} (rol={user.rol}) sin area_codigo → "
                f"visibilidad bloqueada. Asignar area_codigo al usuario."
            )
            return []
        return AREA_JERARQUIA.get(area, [area])

    # Rol desconocido → bloquear por defecto
    logger.warning(f"Rol no reconocido: {user.rol} — bloqueando visibilidad.")
    return []


def _check_area_access(user: Usuario, doc) -> None:
    """Verifica que el usuario tenga permiso de ver un documento específico por área.

    Lanza ForbiddenError si no tiene acceso.
    Usar en endpoints que obtienen un documento por ID directo.

    Reglas:
      - Roles con visión total (Director, Secretaría, Superadmin, Asesor,
        Auditor, Consulta): ven todo.
      - Roles por área (Subdirector, Jefe de Depto, Analista): solo ven
        documentos cuyo area_turno esté en su jerarquía.
      - Documentos sin area_turno asignado (estado 'recibido') son triaje
        inicial — solo los roles con visión total los ven. Los usuarios de
        área NO ven documentos sin turnar.
    """
    areas_vis = _areas_visibles(user)
    if areas_vis is None:
        return  # ve todo
    if not areas_vis:
        raise ForbiddenError("No tiene acceso a documentos (área no asignada).")
    doc_area = getattr(doc, 'area_turno', None)
    # Sin área turnada → documentación de triaje, no es para áreas operativas.
    # Se bloquea explícitamente para subdirector/jefe_depto/analista.
    if not doc_area:
        raise ForbiddenError(
            "Este documento aún no ha sido turnado a un área. "
            "Solo Secretaría y Dirección pueden verlo durante el triaje."
        )
    if doc_area not in areas_vis:
        raise ForbiddenError(
            f"No tiene acceso a documentos del área '{doc_area}'. "
            f"Áreas permitidas: {', '.join(areas_vis)}"
        )


# ---------- Listado ----------------------------------------------------------

@router.get("/", response_model=List[DocumentoListResponse], summary="Listar documentos")
async def listar_documentos(
    response: Response,
    skip:       int  = Query(0, ge=0),
    limit:      int  = Query(200, ge=1, le=500),
    flujo:      Optional[str] = Query(None, description="recibido | emitido"),
    tipo:       Optional[str] = Query(None),
    estado:     Optional[str] = Query(None),
    area_turno: Optional[str] = Query(None),
    cliente_id: Optional[str] = Query(None),
    busqueda:   Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None, description="Fecha inicio YYYY-MM-DD"),
    fecha_hasta: Optional[str] = Query(None, description="Fecha fin YYYY-MM-DD"),
    solo_urgentes: bool = Query(False, description="Solo documentos urgentes o muy urgentes"),
    incluir_respuestas: bool = Query(False, description="Incluir recibidos con oficio de respuesta generado (para tab Documentos emitidos)"),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    if current_user.rol != "superadmin":
        cliente_id = str(current_user.cliente_id)

    # Filtro por estructura organizacional
    areas_vis = _areas_visibles(current_user)
    # Si el usuario pidió filtro manual de area_turno, respetar solo si está en sus visibles
    if area_turno and areas_vis and area_turno not in areas_vis:
        area_turno = None  # no tiene acceso a esa área

    filter_args = dict(
        cliente_id=cliente_id, flujo=flujo, tipo=tipo, estado=estado,
        area_turno=area_turno,
        area_turno_in=areas_vis if not area_turno else None,
        busqueda=busqueda, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
        solo_urgentes=solo_urgentes,
        incluir_respuestas=incluir_respuestas,
    )

    # Obtener total para paginación (header X-Total-Count)
    total = await crud_documento.count_documentos(db, **filter_args)
    response.headers["X-Total-Count"] = str(total)
    response.headers["Access-Control-Expose-Headers"] = "X-Total-Count"

    return await crud_documento.list_documentos(db, **filter_args, skip=skip, limit=limit)


# ---------- Membrete institucional ------------------------------------------

_LOGOS_DIR = os.path.join(
    os.path.dirname(__file__), "..", "..", "..", "static", "logos"
)
os.makedirs(_LOGOS_DIR, exist_ok=True)

_MEMBRETE_MIMES = {"image/png", "image/jpeg", "image/jpg"}
_MEMBRETE_EXTS  = {".png", ".jpg", ".jpeg"}


@router.post("/membrete", summary="Subir membrete institucional activo")
async def subir_membrete(
    file: UploadFile = File(...),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Sube un PNG/JPG que se usará como fondo de página en todos los oficios PDF."""
    if current_user.rol not in ("superadmin", "admin_cliente"):
        raise ForbiddenError("Solo administradores pueden cambiar el membrete.")

    content_type = (file.content_type or "").split(";")[0].strip()
    ext = os.path.splitext(file.filename or "")[1].lower()
    if content_type not in _MEMBRETE_MIMES and ext not in _MEMBRETE_EXTS:
        raise BusinessError("Solo se aceptan imágenes PNG o JPG para el membrete.")

    # Borrar membrete anterior (cualquier extensión)
    for old_ext in _MEMBRETE_EXTS:
        old_path = os.path.join(_LOGOS_DIR, f"membrete_activo{old_ext}")
        if os.path.exists(old_path):
            os.remove(old_path)

    # Guardar nuevo membrete
    save_ext = ext if ext in _MEMBRETE_EXTS else ".png"
    save_path = os.path.join(_LOGOS_DIR, f"membrete_activo{save_ext}")
    content = await file.read()
    with open(save_path, "wb") as f:
        f.write(content)

    return {
        "ok": True,
        "filename": f"membrete_activo{save_ext}",
        "size_kb": round(len(content) / 1024, 1),
        "mensaje": "Membrete actualizado. Los nuevos oficios usarán este fondo.",
    }


@router.get("/membrete/info", summary="Info del membrete activo")
async def info_membrete(
    current_user: Usuario = Depends(get_current_active_user),
):
    """Devuelve información del membrete actualmente activo."""
    for ext in (".png", ".jpg", ".jpeg"):
        path = os.path.join(_LOGOS_DIR, f"membrete_activo{ext}")
        if os.path.exists(path):
            size_kb = round(os.path.getsize(path) / 1024, 1)
            import time
            mtime = time.strftime(
                "%Y-%m-%d %H:%M", time.localtime(os.path.getmtime(path))
            )
            return {
                "activo": True,
                "filename": f"membrete_activo{ext}",
                "size_kb": size_kb,
                "actualizado": mtime,
                "url": f"/api/v1/documentos/membrete/preview",
            }
    return {"activo": False}


@router.get("/membrete/preview", summary="Ver imagen del membrete activo")
async def preview_membrete(
    current_user: Usuario = Depends(get_current_active_user),
):
    """Devuelve la imagen del membrete activo para previsualizarla."""
    for ext in (".png", ".jpg", ".jpeg"):
        path = os.path.join(_LOGOS_DIR, f"membrete_activo{ext}")
        if os.path.exists(path):
            mime = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"
            return FileResponse(path, media_type=mime)
    raise NotFoundError("No hay membrete activo configurado.")


@router.get("/membrete/config", summary="Obtener configuración de coordenadas del membrete")
async def get_membrete_config(
    current_user: Usuario = Depends(get_current_active_user),
):
    """Devuelve la configuración actual de coordenadas y tipografía del membrete."""
    from app.services.oficio_pdf_service import _get_membrete_config
    return _get_membrete_config()


@router.put("/membrete/config", summary="Guardar configuración de coordenadas del membrete")
async def save_membrete_config(
    payload: dict = Body(...),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Persiste la configuración de coordenadas y tipografía del membrete."""
    if current_user.rol not in ("superadmin", "admin_cliente"):
        raise ForbiddenError("Solo administradores pueden modificar la configuración del membrete.")
    from app.services.oficio_pdf_service import _save_membrete_config, _MEMBRETE_CONFIG_DEFAULT
    # Validar que las claves requeridas existan
    required = {"fontsize", "max_chars", "line_height", "fecha_y", "campos"}
    if not required.issubset(payload.keys()):
        raise BusinessError(f"Faltan campos requeridos: {required - payload.keys()}")
    _save_membrete_config(payload)
    return {"ok": True, "mensaje": "Configuración de membrete guardada correctamente."}


@router.get("/membrete/calibrar", summary="PDF de calibración de coordenadas del membrete")
async def calibrar_membrete(
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Genera un PDF con el membrete de fondo + cuadrícula de coordenadas.
    Úsalo para identificar las coordenadas exactas (x, y) de cada campo del recuadro.
    """
    import io as _io
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import letter as _letter
    from reportlab.lib import colors as rl_colors
    from app.services.oficio_pdf_service import (
        _get_membrete_activo, _get_membrete_config
    )

    membrete_path = _get_membrete_activo()
    if not membrete_path:
        raise NotFoundError("No hay membrete activo configurado.")

    _cfg = _get_membrete_config()
    MEMBRETE_CAMPOS = _cfg["campos"]
    MEMBRETE_FECHA_Y = _cfg["fecha_y"]

    buf = _io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=_letter)
    pw, ph = _letter  # 612 × 792 pts

    # 1) Fondo membrete
    c.drawImage(membrete_path, 0, 0, width=pw, height=ph,
                preserveAspectRatio=False, mask="auto")

    # 2) Cuadrícula cada 50 pts
    c.setStrokeColor(rl_colors.HexColor("#AAAAAA"))
    c.setLineWidth(0.3)
    for x in range(0, int(pw) + 1, 50):
        c.line(x, 0, x, ph)
        c.setFont("Helvetica", 5)
        c.setFillColor(rl_colors.HexColor("#666666"))
        c.drawString(x + 1, 2, str(x))
    for y in range(0, int(ph) + 1, 50):
        c.line(0, y, pw, y)
        c.setFont("Helvetica", 5)
        c.setFillColor(rl_colors.HexColor("#666666"))
        c.drawString(2, y + 1, str(y))

    # 3) Marcar posición de cada campo con cruz roja + etiqueta
    nombres = ["Dependencia", "Sub-dep.", "Oficina", "No.oficio", "Expediente", "Asunto"]
    c.setFont("Helvetica-Bold", 7)
    for campo, nombre in zip(MEMBRETE_CAMPOS, nombres):
        x, y = campo["x"], campo["y"]
        c.setStrokeColor(rl_colors.red)
        c.setLineWidth(1)
        c.line(x - 4, y, x + 4, y)
        c.line(x, y - 4, x, y + 4)
        c.setFillColor(rl_colors.HexColor("#0000CC"))
        c.drawString(x + 6, y - 2, f"{nombre} → x={x}, y={y}")

    # 4) Fecha
    c.setStrokeColor(rl_colors.blue)
    c.line(pw - 0.87*72 - 4, MEMBRETE_FECHA_Y, pw - 0.87*72 + 4, MEMBRETE_FECHA_Y)
    c.setFillColor(rl_colors.HexColor("#006600"))
    c.drawString(pw - 0.87*72 - 100, MEMBRETE_FECHA_Y - 8, f"Fecha → y={MEMBRETE_FECHA_Y}")

    c.save()
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": "inline; filename=calibracion_membrete.pdf"})


# ---------- Siguiente folio consecutivo -------------------------------------

@router.get("/siguiente-folio", summary="Obtener siguiente folio consecutivo")
async def siguiente_folio(
    tipo: str = Query("OFICIO", description="Tipo de documento (legacy)"),
    area_codigo: Optional[str] = Query(
        None,
        description="Código del área de origen (DIR, SCG, SPF...). "
                    "Si se proporciona, genera folio institucional SFA/SF/DPP[/AREA]/XXXX/AÑO.",
    ),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Genera el siguiente folio consecutivo para oficios.

    Formato institucional: `SFA/SF/DPP[/AREA]/XXXX/AÑO`
    Ejemplo: `SFA/SF/DPP/SPFP/0001/2026`

    Donde:
      - SFA = Secretaría de Finanzas y Administración
      - SF  = Subsecretaría de Finanzas
      - DPP = Dirección de Programación y Presupuesto
      - AREA = área emisora interna (SPFP, etc.) — opcional según prefijo
      - XXXX = número consecutivo (4 dígitos)
      - AÑO  = año en curso

    Sin area_codigo: formato legacy `DPP/{TIPO}/{NNNNN}/{AÑO}`.
    """
    from datetime import datetime
    from sqlalchemy import text

    anio = datetime.now().year

    def _extraer_numero(folio: str) -> int:
        """Extrae el número consecutivo del penúltimo segmento del folio."""
        try:
            parts = folio.split("/")
            return int(parts[-2])
        except (ValueError, IndexError):
            return 0

    if area_codigo and area_codigo.upper() in PREFIJOS_FOLIO:
        # ── Formato institucional por área ──────────────────────────────
        area = area_codigo.upper()
        prefijo = PREFIJOS_FOLIO[area]
        pattern = f"{prefijo}/%/{anio}"

        # Solo folio_respuesta confirmados (SIN numero_control).
        # Se excluye estado 'de_conocimiento': esos documentos solo acusan
        # recibo y nunca generan un oficio de respuesta; si tienen folio_respuesta
        # es un dato residual que no debe contar.
        # Se excluye estado 'borrador': el folio en un borrador sin confirmar puede
        # haber sido asignado por error y se borrará si el borrador es eliminado.
        query = text(
            "SELECT folio_respuesta FROM documentos_oficiales "
            "WHERE folio_respuesta LIKE :pattern "
            "AND estado NOT IN ('de_conocimiento', 'borrador')"
        )
        result = await db.execute(query, {"pattern": pattern})
        rows = result.fetchall()

        # El folio correcto para este prefijo tiene exactamente:
        #   len(prefijo.split('/')) partes del prefijo + número + año = len+2 segmentos
        # Ejemplo DIR: "SFA/SF/DPP" → 3 segmentos → folio esperado "SFA/SF/DPP/NNNN/YYYY" = 5 segs
        # Esto excluye folios de otras áreas como "SFA/SF/DPP/SCEG/NNNN/YYYY" (6 segs)
        expected_segs = len(prefijo.split("/")) + 2

        next_num = 1
        if rows:
            max_num = max(
                (
                    _extraer_numero(r[0])
                    for r in rows
                    if r[0] and len(r[0].split("/")) == expected_segs
                ),
                default=0,
            )
            next_num = max_num + 1

        folio = f"{prefijo}/{str(next_num).zfill(4)}/{anio}"
        return {
            "folio": folio,
            "numero": next_num,
            "area_codigo": area,
            "prefijo": prefijo,
            "anio": anio,
        }

    # ── Formato legacy (sin área) ───────────────────────────────────────
    tipo_upper = tipo.upper()[:8]
    prefix = f"DPP/{tipo_upper}/"
    suffix = f"/{anio}"
    pattern_legacy = f"{prefix}%{suffix}"

    # Solo folio_respuesta confirmados (sin numero_control externos ni
    # estados de_conocimiento/borrador que no representan folios reales)
    query = text(
        "SELECT folio_respuesta FROM documentos_oficiales "
        "WHERE folio_respuesta LIKE :pattern "
        "AND estado NOT IN ('de_conocimiento', 'borrador')"
    )
    result = await db.execute(query, {"pattern": pattern_legacy})
    rows = result.fetchall()

    next_num = 1
    if rows:
        def _num_legacy(folio: str) -> int:
            try:
                return int(folio.split("/")[2])
            except (ValueError, IndexError):
                return 0
        max_num = max((_num_legacy(r[0]) for r in rows if r[0]), default=0)
        next_num = max_num + 1

    folio = f"{prefix}{str(next_num).zfill(5)}{suffix}"
    return {"folio": folio, "numero": next_num, "tipo": tipo_upper, "anio": anio}


# ---------- Verificar disponibilidad de folio --------------------------------

@router.get("/verificar-folio", summary="Verificar si un folio/No. de oficio ya existe")
async def verificar_folio(
    folio: str = Query(..., description="Folio a verificar, ej: SFA/SF/DPP-SCEG/0001/2026"),
    exclude_id: str = Query(None, description="ID del documento actual (para excluirlo de la búsqueda)"),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Verifica si un número de oficio/folio ya está registrado en el sistema.

    Hace dos verificaciones:
    1. Coincidencia exacta de folio_respuesta o numero_control.
    2. Coincidencia por consecutivo+año: extrae el número de 3-5 dígitos y el
       año del folio (p.ej. "1221" y "2026" de "SFA/SF/DPP/1221/2026") y busca
       cualquier documento cuyo folio_respuesta o numero_control contenga
       ese consecutivo y año — sin importar el prefijo institucional.

    Retorna { disponible: false, documento_id, folio_existente } si hay conflicto.
    """
    import re as _re
    from sqlalchemy import text as sql_text

    folio_limpio = folio.strip()
    exclude = exclude_id.strip() if exclude_id else None

    # ── 0. Validar formato: consecutivo debe ser exactamente 4 dígitos ────────
    m_fmt = _re.search(r'/(\d+)/(20\d{2})$', folio_limpio)
    if m_fmt and len(m_fmt.group(1)) != 4:
        return {
            "disponible": False,
            "folio": folio_limpio,
            "error": "formato",
            "mensaje": f"El consecutivo debe tener exactamente 4 dígitos (recibido: '{m_fmt.group(1)}'). Ejemplo: SFA/SF/DPP/0099/2026",
        }

    # ── 1. Coincidencia exacta ────────────────────────────────────────────────
    excl_clause = "AND id != :excl" if exclude else ""
    result = await db.execute(
        sql_text(
            f"SELECT id, folio_respuesta FROM documentos_oficiales "
            f"WHERE (folio_respuesta = :f OR numero_control = :f) {excl_clause} LIMIT 1"
        ),
        {"f": folio_limpio, **({"excl": exclude} if exclude else {})},
    )
    row = result.first()
    if row:
        return {"disponible": False, "documento_id": str(row[0]),
                "folio_existente": row[1], "folio": folio_limpio}

    # ── 2. Coincidencia por consecutivo + año (sin importar prefijo ni ceros) ──
    # Extrae el número consecutivo (sin ceros) y el año del final del folio.
    # Ej: "SFA/SF/DPP/01221/2026" → consecutivo="1221", anio="2026"
    # Ej: "SFA/SF/DPP/1221/2026"  → consecutivo="1221", anio="2026"
    # Regex PostgreSQL /0*1221/2026$ detecta ambos como duplicados.
    m = _re.search(r'/0*(\d{1,5})/(20\d{2})$', folio_limpio)
    if m:
        # Quitar todos los ceros a la izquierda del consecutivo capturado
        consecutivo_raw = m.group(1).lstrip('0') or '0'
        anio = m.group(2)
        # Patrón: termina en /0*<consecutivo>/<año>  (cualquier cantidad de ceros de padding)
        patron_regex = f"/0*{consecutivo_raw}/{anio}$"

        # Revisar tanto folio_respuesta como numero_control (igual que en paso 1)
        result2 = await db.execute(
            sql_text(
                f"SELECT id, folio_respuesta, numero_control FROM documentos_oficiales "
                f"WHERE (folio_respuesta ~ :patron OR numero_control ~ :patron) "
                f"{excl_clause} LIMIT 1"
            ),
            {"patron": patron_regex, **({"excl": exclude} if exclude else {})},
        )
        row2 = result2.first()
        if row2:
            folio_dup = row2[1] or row2[2]  # folio_respuesta o numero_control
            return {"disponible": False, "documento_id": str(row2[0]),
                    "folio_existente": folio_dup, "folio": folio_limpio,
                    "conflicto": "consecutivo_duplicado"}

    return {"disponible": True, "folio": folio_limpio}


@router.get("/verificar-coherencia-fecha", summary="Verifica coherencia cronológica entre folio y fecha")
async def verificar_coherencia_fecha(
    folio: str = Query(..., description="Folio del documento, ej: SFA/SF/DPP/1263/2026"),
    fecha: str = Query(..., description="Fecha del oficio en español, ej: 20 de abril de 2026"),
    exclude_id: str = Query(None, description="ID del documento actual (excluirlo de la búsqueda)"),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Verifica que un folio con consecutivo N tenga una fecha >= a todos los documentos
    con consecutivo < N del mismo año, y fecha <= a los de consecutivo > N.

    Retorna { coherente: false, conflictos: [...] } si hay incoherencia cronológica.
    """
    import re as _re
    from datetime import date as _date
    from sqlalchemy import text as sql_text

    MESES_ES = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
        'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
        'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
    }

    def _parse_fecha(texto: str):
        m = _re.match(r'^(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})$', texto.strip(), _re.IGNORECASE)
        if not m:
            return None
        mes = MESES_ES.get(m.group(2).lower())
        if not mes:
            return None
        try:
            return _date(int(m.group(3)), mes, int(m.group(1)))
        except ValueError:
            return None

    # Parsear folio actual
    mf = _re.search(r'/0*(\d+)/(20\d{2})$', folio.strip())
    if not mf:
        return {"coherente": True}

    consecutivo_actual = int(mf.group(1))
    anio = mf.group(2)
    fecha_actual = _parse_fecha(fecha)
    if not fecha_actual:
        return {"coherente": True}  # No se puede parsear → no bloquear

    excl_clause = "AND id != :excl" if exclude_id else ""
    params: dict = {"anio_patron": f"%/{anio}", **({"excl": exclude_id} if exclude_id else {})}

    # Incluir TODOS los documentos con folio en el año (con o sin fecha_respuesta).
    # Cuando fecha_respuesta está vacía usamos creado_en como fecha mínima proxy:
    # un documento no puede tener una respuesta fechada ANTES de que entrara al sistema.
    result = await db.execute(
        sql_text(
            f"SELECT id, folio_respuesta, fecha_respuesta, creado_en FROM documentos_oficiales "
            f"WHERE folio_respuesta LIKE :anio_patron "
            f"{excl_clause}"
        ),
        params,
    )
    rows = result.fetchall()

    conflictos = []
    for row in rows:
        otro_folio    = row[1] or ''
        fecha_resp_db = row[2] or ''
        creado_en_db  = row[3]   # ISO datetime string ej: "2026-04-29T..."

        mf2 = _re.search(r'/0*(\d+)/(20\d{2})$', otro_folio)
        if not mf2:
            continue
        otro_consec = int(mf2.group(1))

        # Intentar parsear fecha_respuesta; si está vacía usar creado_en como proxy
        otra_fecha = _parse_fecha(fecha_resp_db) if fecha_resp_db.strip() else None
        fecha_estimada = False
        if otra_fecha is None and creado_en_db:
            # creado_en puede ser "2026-04-29" o "2026-04-29T10:30:00..."
            creado_str = str(creado_en_db)[:10]
            try:
                from datetime import date as _date2
                partes_iso = creado_str.split('-')
                otra_fecha = _date2(int(partes_iso[0]), int(partes_iso[1]), int(partes_iso[2]))
                fecha_estimada = True
            except Exception:
                pass

        if otra_fecha is None:
            continue

        # Nuestro consecutivo > otro → nuestra fecha debe ser >= otra fecha
        if consecutivo_actual > otro_consec and fecha_actual < otra_fecha:
            conflictos.append({
                "folio": otro_folio,
                "fecha": fecha_resp_db if fecha_resp_db.strip() else f"~{str(otra_fecha)} (fecha estimada por creación)",
                "consecutivo": otro_consec,
                "fecha_estimada": fecha_estimada,
                "razon": f"El folio {folio} (consecutivo {consecutivo_actual}) tiene fecha anterior a {otro_folio} (consecutivo {otro_consec})",
            })
        # Nuestro consecutivo < otro → nuestra fecha debe ser <= otra fecha
        elif consecutivo_actual < otro_consec and fecha_actual > otra_fecha:
            conflictos.append({
                "folio": otro_folio,
                "fecha": fecha_resp_db if fecha_resp_db.strip() else f"~{str(otra_fecha)} (fecha estimada por creación)",
                "consecutivo": otro_consec,
                "fecha_estimada": fecha_estimada,
                "razon": f"El folio {folio} (consecutivo {consecutivo_actual}) tiene fecha posterior a {otro_folio} (consecutivo {otro_consec})",
            })

    if conflictos:
        # Ordenar por consecutivo para mostrar el conflicto más cercano primero
        conflictos.sort(key=lambda x: abs(x["consecutivo"] - consecutivo_actual))
        return {"coherente": False, "conflictos": conflictos}

    return {"coherente": True}


# ---------- Catalogo de areas ------------------------------------------------

@router.get("/areas", summary="Catalogo de areas DPP para turno")
async def listar_areas(
    current_user: Usuario = Depends(get_current_active_user),
):
    # Las áreas con asignación restringida (p.ej. SEC) solo se devuelven a los
    # roles autorizados — así el dropdown de "turnar a…" no le muestra la opción
    # a subdirectores/jefes (refuerzo UX del guard del servidor).
    resultado = []
    for k, v in AREAS_DPP.items():
        restric = AREAS_ASIGNACION_RESTRINGIDA.get(k)
        if restric and current_user.rol not in restric["roles_permitidos"]:
            continue
        resultado.append({"codigo": k, **v})
    return resultado


# ---------- Catalogo de plantillas para emitidos ----------------------------

@router.get("/plantillas", summary="Catalogo de plantillas de oficios emitidos")
async def listar_plantillas(
    area_codigo: Optional[str] = Query(None, description="Filtrar por area de origen"),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Retorna las plantillas de oficios disponibles, opcionalmente filtradas por área."""
    result = []
    for p in PLANTILLAS_OFICIO:
        if area_codigo and p["area_origen"] != area_codigo.upper():
            continue
        result.append({
            "categoria": p["categoria"],
            "nombre": p["nombre"],
            "area_origen": p["area_origen"],
            "fundamento_legal": p["fundamento_legal"],
        })
    return result


# ---------- Export Excel oficios recibidos ------------------------------------

@router.get("/export-recibidos", summary="Exportar oficios recibidos a Excel")
async def exportar_recibidos(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    dependencia: Optional[str] = Query(None),
    busqueda: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Genera y descarga un archivo Excel con los oficios recibidos filtrados."""
    import io
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    # Filtrar documentos recibidos
    cliente_id = None if current_user.rol == "superadmin" else str(current_user.cliente_id)
    # Búsqueda por dependencia: se inyecta en busqueda general ya que el CRUD
    # busca en remitente_dependencia entre otros campos
    busqueda_final = busqueda
    if dependencia and busqueda:
        busqueda_final = f"{busqueda} {dependencia}"
    elif dependencia:
        busqueda_final = dependencia

    docs = await crud_documento.list_documentos(
        db,
        cliente_id=cliente_id,
        flujo="recibido",
        busqueda=busqueda_final,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        skip=0,
        limit=10000,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Oficios Recibidos"

    # Estilos
    guinda_fill = PatternFill("solid", fgColor="911A3A")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="top", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    alt_fill = PatternFill("solid", fgColor="FDF2F4")

    # Título
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "Control de Oficios Recibidos — PIGOP"
    title_cell.font = Font(bold=True, size=14, color="911A3A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtítulo con rango de fechas
    ws.merge_cells("A2:I2")
    rango = "Todos los registros"
    if fecha_desde and fecha_hasta:
        rango = f"Del {fecha_desde} al {fecha_hasta}"
    elif fecha_desde:
        rango = f"Desde {fecha_desde}"
    elif fecha_hasta:
        rango = f"Hasta {fecha_hasta}"
    ws["A2"].value = rango
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Encabezados
    columns = [
        ("Folio", 8),
        ("No. Oficio", 28),
        ("Asunto", 40),
        ("Remitente", 25),
        ("Dependencia", 30),
        ("Estado", 14),
        ("Prioridad", 12),
        ("Fecha Oficio", 14),
        ("Fecha Recibido", 14),
    ]

    header_row = 4
    for col_idx, (col_name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = guinda_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = width

    # Datos
    for row_idx, doc in enumerate(docs, start=header_row + 1):
        folio = row_idx - header_row
        values = [
            folio,
            doc.numero_oficio_origen or "",
            doc.asunto or "",
            doc.remitente_nombre or "",
            doc.remitente_dependencia or doc.dependencia_origen or "",
            doc.estado or "",
            doc.prioridad or "normal",
            doc.fecha_documento or "",
            doc.fecha_recibido or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin
            cell.alignment = wrap
            if (row_idx - header_row) % 2 == 0:
                cell.fill = alt_fill

    # Footer
    footer_row = header_row + len(docs) + 2
    ws.merge_cells(f"A{footer_row}:I{footer_row}")
    ws[f"A{footer_row}"].value = f"Total de oficios: {len(docs)}"
    ws[f"A{footer_row}"].font = Font(bold=True, size=10, color="911A3A")

    # Serializar
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = "PIGOP_Control_Oficios"
    if fecha_desde:
        filename += f"_{fecha_desde}"
    if fecha_hasta:
        filename += f"_{fecha_hasta}"
    filename += ".xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export-emitidos", summary="Exportar documentos emitidos a Excel")
async def exportar_emitidos(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    busqueda: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Genera y descarga un archivo Excel con los documentos emitidos."""
    import io
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    cliente_id = None if current_user.rol == "superadmin" else str(current_user.cliente_id)
    docs = await crud_documento.list_documentos(
        db, cliente_id=cliente_id, flujo="emitido", busqueda=busqueda,
        fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, skip=0, limit=10000,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Documentos Emitidos"

    guinda_fill = PatternFill("solid", fgColor="911A3A")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="top", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["No. Oficio", "Asunto", "Área", "Destinatario", "Estado", "Fecha", "UPP"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = guinda_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    for doc in docs:
        ws.append([
            doc.folio_respuesta or doc.numero_control or "",
            doc.asunto or "",
            doc.area_turno_nombre or "",
            doc.destinatario_nombre or doc.dependencia_destino or "",
            doc.estado or "",
            doc.fecha_documento or "",
            doc.upp_solicitante or "",
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap
            cell.border = thin

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    filename = f"emitidos_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------- Crear recibido ---------------------------------------------------

@router.post(
    "/recibido",
    response_model=DocumentoResponse,
    status_code=201,
    summary="Registrar oficio recibido",
)
async def crear_recibido(
    data: DocumentoRecibidoCreate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Registra un oficio recibido en la DPP.
    Acepta datos pre-llenados por preview-ocr (archivo, OCR, clasificacion).
    """
    _assert_acceso(current_user, data.cliente_id)
    doc = await crud_documento.crear_recibido(
        db, obj_in=data, creado_por_id=str(current_user.id)
    )
    return await crud_documento.get_with_relations(db, doc.id)


# ---------- Registrar memorándum ---------------------------------------------

@router.post(
    "/memorandum",
    response_model=DocumentoResponse,
    status_code=201,
    summary="Registrar memorándum institucional",
)
async def registrar_memorandum(
    data: DocumentoRecibidoCreate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Registra un memorándum institucional proveniente de Secretaría o Subsecretaría de Finanzas.

    Reglas de negocio:
    - Emisor: siempre Secretaría o Subsecretaría de Finanzas
    - tipo_memorandum='conocimiento' → estado directo 'de_conocimiento'
    - tipo_memorandum='requiere_atencion' → estado 'recibido' (flujo normal)
    - La respuesta va a la dependencia_solicitante, NO al emisor del memo
    - memorandum_orden_direccion=1 → esta dirección es responsable
    - memorandum_orden_direccion>1 → solo conocimiento
    """
    _assert_acceso(current_user, data.cliente_id)

    # Forzar tipo memorandum y flujo recibido
    data.tipo = "memorandum"

    # Si orden > 1, forzar conocimiento
    if data.memorandum_orden_direccion and data.memorandum_orden_direccion > 1:
        data.tipo_memorandum = "conocimiento"
        data.requiere_respuesta = False

    # Si tipo_memorandum es conocimiento, forzar no requiere respuesta
    if data.tipo_memorandum == "conocimiento":
        data.requiere_respuesta = False

    doc = await crud_documento.crear_recibido(
        db, obj_in=data, creado_por_id=str(current_user.id)
    )
    return await crud_documento.get_with_relations(db, doc.id)


# ---------- Crear emitido ----------------------------------------------------

@router.post(
    "/emitido",
    response_model=DocumentoResponse,
    status_code=201,
    summary="Registrar documento emitido por DPP",
)
async def crear_emitido(
    data: DocumentoEmitidoCreate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    _assert_acceso(current_user, data.cliente_id)
    doc = await crud_documento.crear_emitido(
        db, obj_in=data, creado_por_id=str(current_user.id)
    )
    return await crud_documento.get_with_relations(db, doc.id)


# ---------- Preview OCR (sin crear documento) --------------------------------

@router.post(
    "/preview-ocr",
    response_model=PreviewOCRResponse,
    summary="Subir scan, procesar OCR y devolver datos extraidos (sin crear documento)",
)
async def preview_ocr(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Primer paso del flujo inteligente:
    1. Recibe scan/foto del oficio
    2. Guarda el archivo
    3. Procesa OCR con Gemini Vision
    4. Clasifica area de turno
    5. Devuelve datos extraidos + info del archivo guardado
    El documento NO se crea todavia -- el frontend muestra los datos
    para que la secretaria revise/corrija antes de confirmar.
    """
    if file.content_type not in MIME_PERMITIDOS:
        raise BusinessError(f"Tipo de archivo no permitido: {file.content_type}")

    content = await file.read()
    if len(content) > MAX_SIZE_MB * 1024 * 1024:
        raise BusinessError(f"El archivo supera el limite de {MAX_SIZE_MB} MB.")

    # Guardar archivo
    ext = os.path.splitext(file.filename or "oficio.pdf")[1] or ".pdf"
    filename = f"{uuid.uuid4()}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, "wb") as f:
        f.write(content)

    # Procesar OCR + clasificar
    resultado = await correspondencia_service.procesar_oficio_escaneado(
        image_bytes=content,
        mime_type=file.content_type or "image/jpeg",
    )

    # ── Detección de duplicados (Punto 10) ──
    numero_oficio_detectado = (resultado["datos_extraidos"] or {}).get("numero_oficio")
    duplicado_encontrado = None
    if numero_oficio_detectado:
        from sqlalchemy import select, func
        from app.models.documento import DocumentoOficial
        cid = str(current_user.cliente_id) if current_user.cliente_id else None
        stmt = select(DocumentoOficial.id, DocumentoOficial.numero_oficio_origen, DocumentoOficial.asunto, DocumentoOficial.fecha_documento).where(
            func.lower(DocumentoOficial.numero_oficio_origen) == numero_oficio_detectado.lower()
        )
        if cid:
            stmt = stmt.where(DocumentoOficial.cliente_id == cid)
        from app.core.database import get_db as _gdb
        dup_result = await db.execute(stmt)
        dup_row = dup_result.first()
        if dup_row:
            duplicado_encontrado = {
                "id": dup_row[0],
                "numero_oficio": dup_row[1],
                "asunto": dup_row[2],
                "fecha": dup_row[3],
            }

    msg = "Datos extraídos correctamente. Revise y confirme."
    if duplicado_encontrado:
        msg = f"⚠️ POSIBLE DUPLICADO: Ya existe un oficio con número '{numero_oficio_detectado}' registrado en el sistema."

    return PreviewOCRResponse(
        datos_extraidos=resultado["datos_extraidos"],
        clasificacion=resultado["clasificacion"],
        fecha_limite=resultado["fecha_limite"],
        archivo={
            "nombre_archivo": file.filename or filename,
            "url_storage": filepath,
            "mime_type": file.content_type or "application/octet-stream",
        },
        message=msg,
        prioridad_sugerida=resultado.get("prioridad_sugerida", "normal"),
        duplicado=duplicado_encontrado,
    )


# ---------- Listar documentos devueltos (ANTES de /{doc_id}) ----------------

@router.get(
    "/devueltos",
    response_model=List[DocumentoListResponse],
    summary="Listar documentos devueltos",
)
async def listar_devueltos(
    area_turno: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Lista documentos en estado 'devuelto' para el área responsable."""
    cliente_id = (
        None if current_user.rol == "superadmin" else str(current_user.cliente_id)
    )
    # Aplicar filtro por áreas visibles (igual que en listar_documentos).
    # Evita que un subdirector/jefe vea devoluciones de otra área.
    areas_vis = _areas_visibles(current_user)
    if area_turno and areas_vis and area_turno not in areas_vis:
        area_turno = None  # sin acceso → no mostrar nada de esa área
    return await crud_documento.list_devueltos(
        db, cliente_id=cliente_id, area_turno=area_turno,
        area_turno_in=areas_vis if not area_turno else None,
    )


# ---------- Obtener uno ------------------------------------------------------

@router.get("/{doc_id}", response_model=DocumentoResponse, summary="Obtener documento")
async def obtener_documento(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    doc = await crud_documento.get_with_relations(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))
    _check_area_access(current_user, doc)
    return doc


# ---------- Actualizar -------------------------------------------------------

@router.put("/{doc_id}", response_model=DocumentoResponse, summary="Actualizar documento")
async def actualizar_documento(
    doc_id: str,
    data: DocumentoUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))
    updated = await crud_documento.actualizar_documento(db, db_obj=doc, obj_in=data)
    return await crud_documento.get_with_relations(db, updated.id)


# ---------- Cambiar estado ---------------------------------------------------

# Mapa de transiciones válidas por flujo
TRANSICIONES_RECIBIDO = {
    "recibido": ["turnado"],
    "turnado": ["en_atencion"],
    "en_atencion": ["respondido", "de_conocimiento"],
    "respondido": ["firmado", "en_atencion"],   # permite revertir si aún no está firmado
    "firmado": ["archivado"],
    "de_conocimiento": ["en_atencion", "respondido", "archivado"],
    "devuelto": ["en_atencion"],
}

TRANSICIONES_EMITIDO = {
    "borrador": ["en_revision"],
    "en_revision": ["vigente", "borrador"],
    "vigente": ["archivado"],
}


@router.post("/{doc_id}/estado", response_model=DocumentoResponse, summary="Cambiar estado")
async def cambiar_estado(
    doc_id: str,
    nuevo_estado: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    validos = ESTADOS_RECIBIDO + ESTADOS_EMITIDO
    if nuevo_estado not in validos:
        raise BusinessError(f"Estado invalido. Opciones: {', '.join(validos)}")
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    # Regla operativa: si requiere_respuesta=True, no se puede pasar a de_conocimiento
    if nuevo_estado == "de_conocimiento" and doc.requiere_respuesta:
        raise BusinessError(
            "Este oficio requiere respuesta obligatoria. "
            "No puede marcarse como 'De conocimiento'. "
            "Debe atenderse: En atención → Respondido → Firmado."
        )

    # Detectar caso idempotente: el documento ya está en el estado destino
    # (p.ej. el usuario reintenta "Enviar a firma" cuando ya se envió antes).
    # En ese caso NO fallamos por transición inválida, pero aún permitimos
    # aplicar side-effects puntuales (como el visto bueno del subdirector).
    es_idempotente = nuevo_estado == doc.estado

    # Validar transición de estado (superadmin puede hacer cualquier transición).
    # La idempotencia se permite aunque "respondido → respondido" no esté en el mapa.
    if current_user.rol != "superadmin" and not es_idempotente:
        mapa = TRANSICIONES_RECIBIDO if doc.flujo == "recibido" else TRANSICIONES_EMITIDO
        destinos_permitidos = mapa.get(doc.estado, [])
        if nuevo_estado not in destinos_permitidos:
            raise BusinessError(
                f"Transición no permitida: '{doc.estado}' → '{nuevo_estado}'. "
                f"Transiciones válidas desde '{doc.estado}': {', '.join(destinos_permitidos) or 'ninguna'}."
            )

    update_payload: dict = {} if es_idempotente else {"estado": nuevo_estado}

    # Auto-check Visto Bueno del Subdirector: si quien envía a firma (estado='respondido')
    # es un subdirector, el visto bueno se marca automáticamente (elaboró y revisó él mismo).
    # Esto ocurre tanto cuando se transiciona a 'respondido' como cuando el documento ya
    # está 'respondido' y el subdirector vuelve a confirmar — así el "check" siempre queda
    # registrado cuando corresponde.
    if (
        nuevo_estado == "respondido"
        and current_user.rol == "subdirector"
        and not doc.visto_bueno_subdirector
    ):
        from datetime import datetime, timezone
        update_payload["visto_bueno_subdirector"] = True
        update_payload["visto_bueno_subdirector_id"] = str(current_user.id)
        update_payload["visto_bueno_subdirector_en"] = datetime.now(timezone.utc)
        logger.info(
            f"[VB-AUTO] Subdirector {current_user.email} envió a firma → "
            f"visto bueno auto-marcado para doc {doc_id[:8]}..."
        )

    # Si es idempotente y no hay side-effects pendientes, no hacemos UPDATE.
    if not update_payload:
        return await crud_documento.get_with_relations(db, doc.id)

    updated = await crud_documento.update(db, db_obj=doc, obj_in=update_payload)
    return await crud_documento.get_with_relations(db, updated.id)


# ---------- Cambiar tipo: respuesta ↔ conocimiento (Secretaría) --------------

@router.post(
    "/{doc_id}/cambiar-tipo-respuesta",
    response_model=DocumentoResponse,
    summary="Secretaría cambia si un oficio requiere respuesta o es solo de conocimiento",
)
async def cambiar_tipo_respuesta(
    doc_id: str,
    requiere_respuesta: bool = Body(..., embed=True, description="True = requiere respuesta, False = solo conocimiento"),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Permite a la Secretaría (o Director/Superadmin) cambiar si un oficio
    requiere respuesta o pasarlo a solo conocimiento.

    Reglas:
    - True → False (requiere respuesta → conocimiento): permitido si no está
      firmado. Cambia el estado a 'de_conocimiento' si estaba en recibido/turnado/en_atencion.
    - False → True (conocimiento → requiere respuesta): permitido si está en
      'de_conocimiento' y no fue archivado. Cambia a 'turnado' o 'en_atencion'.
    - No aplica a documentos emitidos (solo flujo recibido).
    """
    if current_user.rol not in ("secretaria", "admin_cliente", "superadmin"):
        raise ForbiddenError(
            "Solo Secretaría, Director o Superadmin pueden cambiar el tipo de oficio."
        )

    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    if doc.flujo != "recibido":
        raise BusinessError("Esta acción solo aplica a oficios recibidos.")

    if doc.firmado_digitalmente or doc.estado in ("firmado", "archivado"):
        raise BusinessError(
            f"No se puede cambiar el tipo de un oficio ya firmado o archivado. Estado: {doc.estado}"
        )

    updates: dict = {"requiere_respuesta": requiere_respuesta}

    # Reglas de transición automática de estado:
    if not requiere_respuesta:
        # → De conocimiento: pasar a estado de_conocimiento si está en flujo de atención
        if doc.estado in ("recibido", "turnado", "en_atencion"):
            updates["estado"] = "de_conocimiento"
    else:
        # → Requiere respuesta: si estaba en de_conocimiento, regresar a turnado/en_atencion
        if doc.estado == "de_conocimiento":
            # Si ya hay área turnada, va a en_atencion; si no, a turnado
            updates["estado"] = "en_atencion" if doc.area_turno else "turnado"

    from app.models.documento import HistorialDocumento

    updated = await crud_documento.update(db, db_obj=doc, obj_in=updates)

    # Registrar en historial
    historial = HistorialDocumento(
        documento_id=str(doc.id),
        usuario_id=str(current_user.id),
        tipo_accion="cambio_tipo_respuesta",
        estado_anterior=doc.estado,
        estado_nuevo=updates.get("estado", doc.estado),
        observaciones=(
            f"Secretaría cambió tipo: "
            f"{'Requiere respuesta' if requiere_respuesta else 'Solo conocimiento'}"
        ),
        version=doc.version or 1,
    )
    db.add(historial)
    await db.commit()

    return await crud_documento.get_with_relations(db, updated.id)


# ---------- Acuse de conocimiento --------------------------------------------

@router.post(
    "/{doc_id}/acusar-conocimiento",
    response_model=DocumentoResponse,
    summary="Marcar documento de conocimiento como atendido",
)
async def acusar_conocimiento(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))
    if doc.estado != "de_conocimiento":
        raise BusinessError("Solo documentos con estado 'de_conocimiento' pueden acusarse como atendidos.")
    area_nombre = doc.area_turno_nombre or "Sin área asignada"
    await crud_documento.acusar_conocimiento(
        db, db_obj=doc, usuario_id=str(current_user.id), area_nombre=area_nombre,
    )
    return await crud_documento.get_with_relations(db, doc.id)


# ---------- Acusar despacho (secretaria) --------------------------------------

@router.post(
    "/{doc_id}/acusar-despacho",
    response_model=DocumentoResponse,
    summary="Secretaria acusa despacho de documento firmado",
)
async def acusar_despacho(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Secretaria marca que el documento firmado ya fue despachado físicamente."""
    if current_user.rol not in ("secretaria", "superadmin"):
        raise ForbiddenError("Solo Secretaría o Administrador pueden acusar despacho.")
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    if not doc.firmado_digitalmente:
        raise BusinessError("Solo documentos firmados pueden acusarse como despachados.")
    from datetime import datetime, timezone
    await crud_documento.update(db, db_obj=doc, obj_in={
        "despachado": True,
        "despachado_por_id": str(current_user.id),
        "despachado_en": datetime.now(timezone.utc),
    })
    return await crud_documento.get_with_relations(db, doc.id)


@router.post(
    "/acusar-despacho-lote",
    response_model=MessageResponse,
    summary="Acusar despacho de múltiples documentos firmados",
)
async def acusar_despacho_lote(
    ids: List[str] = Body(..., embed=True),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Secretaria marca múltiples documentos firmados como despachados."""
    if current_user.rol not in ("secretaria", "superadmin"):
        raise ForbiddenError("Solo Secretaría o Administrador pueden acusar despacho.")
    from datetime import datetime, timezone
    count = 0
    for doc_id in ids:
        doc = await crud_documento.get(db, doc_id)
        if doc and doc.firmado_digitalmente and not doc.despachado:
            await crud_documento.update(db, db_obj=doc, obj_in={
                "despachado": True,
                "despachado_por_id": str(current_user.id),
                "despachado_en": datetime.now(timezone.utc),
            })
            count += 1
    return MessageResponse(message=f"{count} documento(s) marcado(s) como despachado(s)", success=True)


# ---------- Visto Bueno del Subdirector --------------------------------------

@router.post(
    "/{doc_id}/visto-bueno",
    response_model=DocumentoResponse,
    summary="Subdirector registra visto bueno",
)
async def registrar_visto_bueno(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    if current_user.rol not in ("subdirector", "admin_cliente", "superadmin"):
        raise ForbiddenError("Solo Subdirectores, Director o superadmin pueden dar visto bueno.")

    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    
    _assert_acceso(current_user, str(doc.cliente_id))

    from datetime import datetime, timezone
    from app.models.documento import HistorialDocumento
    try:
        to_upd = {"visto_bueno_subdirector": True}
        if hasattr(doc, 'visto_bueno_subdirector_id'):
            to_upd["visto_bueno_subdirector_id"] = str(current_user.id)
        if hasattr(doc, 'visto_bueno_subdirector_en'):
            to_upd["visto_bueno_subdirector_en"] = datetime.now(timezone.utc)
            
        await crud_documento.update(db, db_obj=doc, obj_in=to_upd)

        try:
            historial = HistorialDocumento(
                documento_id=doc.id,
                usuario_id=str(current_user.id),
                tipo_accion="visto_bueno",
                estado_anterior=doc.estado,
                estado_nuevo=doc.estado,
                observaciones=f"Visto Bueno registrado por {current_user.nombre_completo or current_user.email}",
                version=doc.version or 1,
            )
            db.add(historial)
        except:
            pass

        await db.commit()
        return await crud_documento.get_with_relations(db, doc_id)
    except Exception as e:
        await db.rollback()
        raise BusinessError(f"Error en base de datos: {str(e)}")


# ---------- Procesar OCR -----------------------------------------------------

@router.post(
    "/{doc_id}/procesar-ocr",
    response_model=ProcesarOCRResponse,
    summary="Procesar scan con OCR + clasificar area de turno",
)
async def procesar_ocr(
    doc_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    1. Recibe el scan/foto del oficio
    2. Guarda el archivo
    3. Envia a Gemini Vision para OCR
    4. Clasifica el asunto -> area de turno
    5. Calcula fecha limite
    6. Guarda todo en el documento
    """
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    if file.content_type not in MIME_PERMITIDOS:
        raise BusinessError(f"Tipo de archivo no permitido: {file.content_type}")

    content = await file.read()
    if len(content) > MAX_SIZE_MB * 1024 * 1024:
        raise BusinessError(f"El archivo supera el limite de {MAX_SIZE_MB} MB.")

    # Guardar archivo
    ext = os.path.splitext(file.filename or "oficio.pdf")[1] or ".pdf"
    filename = f"{uuid.uuid4()}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, "wb") as f:
        f.write(content)

    # Actualizar archivo en el registro
    await crud_documento.actualizar_archivo(
        db,
        db_obj=doc,
        nombre_archivo=file.filename or filename,
        url_storage=filepath,
        mime_type=file.content_type or "application/octet-stream",
    )
    # Refrescar
    doc = await crud_documento.get(db, doc_id)

    # Procesar con Gemini Vision + clasificar
    resultado = await correspondencia_service.procesar_oficio_escaneado(
        image_bytes=content,
        mime_type=file.content_type or "image/jpeg",
    )

    # Guardar resultados
    await crud_documento.registrar_ocr(
        db,
        db_obj=doc,
        datos_extraidos=resultado["datos_extraidos"],
        clasificacion=resultado["clasificacion"],
        fecha_limite=resultado["fecha_limite"],
    )

    return ProcesarOCRResponse(
        datos_extraidos=resultado["datos_extraidos"],
        clasificacion=resultado["clasificacion"],
        fecha_limite=resultado["fecha_limite"],
        message="OCR procesado. Revise los datos extraidos y confirme el area de turno.",
    )


# ---------- Clasificar por asunto (sin archivo) ------------------------------

@router.post(
    "/{doc_id}/clasificar",
    response_model=DocumentoResponse,
    summary="Clasificar area de turno por asunto (sin OCR)",
)
async def clasificar_por_asunto(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Clasifica el area de turno usando solo el asunto ya registrado."""
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    from app.services.correspondencia_service import calcular_fecha_limite
    from datetime import date

    clasificacion = clasificar_oficio(doc.asunto or "", "")
    fecha_limite = calcular_fecha_limite(date.today(), clasificacion["plazo_dias"])

    await crud_documento.registrar_ocr(
        db,
        db_obj=doc,
        datos_extraidos=doc.datos_extraidos_ia or {},
        clasificacion=clasificacion,
        fecha_limite=fecha_limite.isoformat(),
    )
    return await crud_documento.get_with_relations(db, doc.id)


# ---------- Confirmar turno --------------------------------------------------

@router.post(
    "/{doc_id}/confirmar-turno",
    response_model=DocumentoResponse,
    summary="Confirmar area de turno (secretaria / habilitado)",
)
async def confirmar_turno(
    doc_id: str,
    data: ConfirmarTurnoInput,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Director, Secretaría o superadmin pueden confirmar el turno."""
    if current_user.rol in ("consulta", "analista"):
        raise ForbiddenError("No tienes permisos para turnar correspondencia.")

    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    if data.area_codigo not in AREAS_DPP:
        raise BusinessError(
            f"Area invalida '{data.area_codigo}'. Validas: {', '.join(AREAS_DPP.keys())}"
        )

    # Guardia: áreas con asignación restringida (p.ej. SEC = Secretaría del Director).
    # Solo los roles permitidos pueden turnar hacia estas bandejas.
    restric = AREAS_ASIGNACION_RESTRINGIDA.get(data.area_codigo)
    if restric and current_user.rol not in restric["roles_permitidos"]:
        raise ForbiddenError(restric["mensaje"])

    area_info = AREAS_DPP[data.area_codigo]
    updated = await crud_documento.confirmar_turno(
        db,
        db_obj=doc,
        area_codigo=data.area_codigo,
        area_nombre=data.area_nombre or area_info["nombre"],
        turnado_por_id=str(current_user.id),
        instrucciones=data.instrucciones,
    )
    return await crud_documento.get_with_relations(db, updated.id)


# ---------- Cambiar turno (re-turnar) -----------------------------------------

@router.post(
    "/{doc_id}/cambiar-turno",
    response_model=DocumentoResponse,
    summary="Cambiar área de turno de un documento ya turnado",
)
async def cambiar_turno(
    doc_id: str,
    data: ConfirmarTurnoInput,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Secretaría, Director, Subdirector, Jefe de Depto o superadmin pueden re-turnar."""
    if current_user.rol in ("consulta", "auditor"):
        raise ForbiddenError("No tienes permisos para cambiar el turno.")

    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    if data.area_codigo not in AREAS_DPP:
        raise BusinessError(
            f"Área inválida '{data.area_codigo}'. Válidas: {', '.join(AREAS_DPP.keys())}"
        )

    # Guardia: áreas con asignación restringida (p.ej. SEC = Secretaría del Director).
    # Evita que subdirectores o jefes redirijan trabajo a la Secretaría del Director.
    restric = AREAS_ASIGNACION_RESTRINGIDA.get(data.area_codigo)
    if restric and current_user.rol not in restric["roles_permitidos"]:
        raise ForbiddenError(restric["mensaje"])

    # Regla adicional: áreas operativas (subdirector, jefe_depto, analista) solo
    # pueden redirigir a su propia subárea o departamentos bajo su línea jerárquica,
    # no a DIR ni a otras subdirecciones. Esto refuerza la compartimentación.
    if current_user.rol in ("subdirector", "analista"):
        user_area = getattr(current_user, "area_codigo", None)
        permitidas = AREA_JERARQUIA.get(user_area, [user_area] if user_area else [])
        if permitidas is None:
            permitidas = []
        if data.area_codigo not in permitidas:
            raise ForbiddenError(
                f"No puedes redirigir a '{data.area_codigo}'. Las áreas permitidas "
                f"para tu rol son: {', '.join(permitidas) or 'ninguna'}. "
                f"Solicita al Director o a la Secretaría reasignar fuera de tu área."
            )

    # Registrar en historial el cambio de turno
    from app.models.documento import HistorialDocumento
    import uuid as _uuid
    historial = HistorialDocumento(
        id=str(_uuid.uuid4()),
        documento_id=doc.id,
        usuario_id=str(current_user.id),
        tipo_accion="cambio_turno",
        estado_anterior=doc.estado,
        estado_nuevo="turnado",
        observaciones=f"Cambio de turno: {doc.area_turno_nombre or '(sin área)'} → {AREAS_DPP[data.area_codigo]['nombre']}. {data.instrucciones or ''}",
        version=doc.version or 1,
    )
    db.add(historial)

    area_info = AREAS_DPP[data.area_codigo]
    updated = await crud_documento.confirmar_turno(
        db,
        db_obj=doc,
        area_codigo=data.area_codigo,
        area_nombre=data.area_nombre or area_info["nombre"],
        turnado_por_id=str(current_user.id),
        instrucciones=data.instrucciones,
    )
    return await crud_documento.get_with_relations(db, updated.id)


# ---------- Generar borrador de respuesta ------------------------------------

@router.post(
    "/{doc_id}/generar-borrador",
    response_model=DocumentoResponse,
    summary="Generar borrador de respuesta con IA",
)
async def generar_borrador(
    doc_id: str,
    body: Optional[dict] = None,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    doc = await crud_documento.get_with_relations(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    instrucciones = ""
    if body and isinstance(body, dict):
        instrucciones = body.get("instrucciones", "")

    area_nombre  = doc.area_turno_nombre or doc.sugerencia_area_nombre or "Direccion de Programacion y Presupuesto"
    fundamento   = doc.sugerencia_fundamento or "Arts. 18 y 19 del Reglamento Interior de la SFA"

    # Contenido de referencia cargado por el usuario (tabla, borrador previo, etc.)
    contenido_ref = doc.contenido_referencia or ""

    # Leer archivo de referencia como bytes para envío multimodal a Gemini
    # NOTA: Solo enviar bytes para tipos que Gemini soporta como multimodal
    # (imágenes y PDF). Para .docx/.xlsx/.csv/.txt usamos el texto extraído
    # que ya está en contenido_referencia.
    ref_bytes: bytes | None = None
    ref_mime: str = ""
    # MIME types que Gemini acepta como entrada multimodal
    GEMINI_MULTIMODAL_MIMES = {
        ".pdf": "application/pdf",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".tiff": "image/tiff",
        ".tif": "image/tiff",
        ".webp": "image/webp",
    }
    if doc.referencia_archivo_url and os.path.exists(doc.referencia_archivo_url):
        nombre = (doc.referencia_archivo_nombre or "").lower()
        # Buscar extensión soportada por Gemini multimodal
        ext = os.path.splitext(nombre)[1]
        if ext in GEMINI_MULTIMODAL_MIMES:
            try:
                with open(doc.referencia_archivo_url, "rb") as rf:
                    ref_bytes = rf.read()
                ref_mime = GEMINI_MULTIMODAL_MIMES[ext]
            except Exception as e:
                logger.warning(f"No se pudo leer archivo de referencia: {e}")
                ref_bytes = None
        else:
            # Para .docx, .xlsx, .csv, .txt etc. → usar solo texto extraído
            # (ya está en contenido_ref desde el proceso de carga)
            logger.info(f"Archivo referencia '{nombre}' no es multimodal Gemini, usando texto extraído.")

    if doc.flujo == "emitido":
        borrador = await correspondencia_service.generar_borrador_emitido(
            asunto=doc.asunto,
            destinatario=doc.dependencia_destino or "",
            dependencia_destino=doc.dependencia_destino or "",
            area_codigo=doc.area_turno or "DIR",
            instrucciones=instrucciones,
            contenido_referencia=contenido_ref,
            referencia_archivo_bytes=ref_bytes,
            referencia_mime_type=ref_mime,
        )
    else:
        resumen_ocr = (doc.datos_extraidos_ia or {}).get("cuerpo_resumen", "")
        borrador = await correspondencia_service.generar_borrador_respuesta(
            numero_oficio_origen=doc.numero_oficio_origen or "---",
            fecha_recibido=doc.fecha_recibido or doc.fecha_documento or "---",
            remitente_nombre=doc.remitente_nombre or "---",
            remitente_cargo=doc.remitente_cargo or "---",
            remitente_dependencia=doc.remitente_dependencia or "---",
            asunto=doc.asunto,
            cuerpo_resumen=resumen_ocr,
            area_nombre=area_nombre,
            fundamento_legal=fundamento,
            instrucciones=instrucciones,
            contenido_referencia=contenido_ref,
            referencia_archivo_bytes=ref_bytes,
            referencia_mime_type=ref_mime,
        )

    updated = await crud_documento.guardar_borrador(db, db_obj=doc, borrador=borrador)
    return await crud_documento.get_with_relations(db, updated.id)


# ---------- Helper: convertir DOCX a PDF usando LibreOffice -------------------

def _convertir_docx_a_pdf_externo(docx_path: str, output_dir: str, doc_id: str) -> str:
    """Convierte un archivo DOCX a PDF usando LibreOffice en modo headless.

    Devuelve la ruta al PDF generado. Si la conversión falla (LibreOffice no
    instalado, error de proceso, etc.) devuelve la ruta original del DOCX como
    fallback para que el flujo no se interrumpa.
    """
    import shutil
    import subprocess

    pdf_target = os.path.join(output_dir, f"oficio_externo_{doc_id}.pdf")

    # Buscar ejecutable de LibreOffice (nombre varía por SO)
    lo_exec = shutil.which("libreoffice") or shutil.which("soffice")
    if lo_exec is None:
        logger.warning("LibreOffice no encontrado; no se puede convertir DOCX → PDF.")
        return docx_path

    try:
        result = subprocess.run(
            [lo_exec, "--headless", "--convert-to", "pdf", "--outdir", output_dir, docx_path],
            capture_output=True,
            timeout=90,
        )
        # LibreOffice genera el PDF con el mismo nombre base del DOCX
        base_name = os.path.splitext(os.path.basename(docx_path))[0]
        generated_pdf = os.path.join(output_dir, f"{base_name}.pdf")

        if result.returncode == 0 and os.path.exists(generated_pdf):
            # Renombrar al nombre canónico para evitar colisiones
            if generated_pdf != pdf_target:
                shutil.move(generated_pdf, pdf_target)
            logger.info("DOCX convertido a PDF: %s → %s", docx_path, pdf_target)
            return pdf_target
        else:
            logger.error(
                "LibreOffice falló (rc=%s): %s",
                result.returncode,
                result.stderr.decode(errors="replace"),
            )
    except subprocess.TimeoutExpired:
        logger.error("LibreOffice excedió el tiempo límite convirtiendo %s", docx_path)
    except Exception as exc:
        logger.error("Error convirtiendo DOCX a PDF: %s", exc)

    # Fallback: devolver el DOCX original
    return docx_path


# ---------- Subir oficio elaborado externamente --------------------------------

@router.post(
    "/{doc_id}/subir-oficio-externo",
    response_model=DocumentoResponse,
    summary="Subir oficio de respuesta elaborado fuera de la plataforma",
)
async def subir_oficio_externo(
    doc_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    doc = await crud_documento.get_with_relations(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    # Guardar el archivo en la carpeta de uploads del documento
    upload_dir = os.path.join("uploads", str(doc.cliente_id), "oficios_externos")
    os.makedirs(upload_dir, exist_ok=True)
    ext = os.path.splitext(file.filename or "oficio.pdf")[1].lower() or ".pdf"
    safe_name = f"oficio_externo_{doc_id}{ext}"
    file_path = os.path.join(upload_dir, safe_name)

    contents = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)

    # ── Si es DOCX, convertir a PDF para que la vista previa lo muestre tal cual
    pdf_path = file_path  # por defecto apunta al mismo archivo si ya es PDF
    if ext in (".docx", ".doc"):
        pdf_path = _convertir_docx_a_pdf_externo(file_path, upload_dir, doc_id)

    # Guardar referencia: oficio_externo_url apunta al PDF (o al original si ya era PDF)
    doc.oficio_externo_url = pdf_path
    doc.oficio_externo_nombre = file.filename or safe_name
    # Mantener el marcador para compatibilidad con otros flujos
    doc.borrador_respuesta = f"[OFICIO EXTERNO: {file.filename}]"

    db.add(doc)
    await db.commit()
    await db.refresh(doc)
    return await crud_documento.get_with_relations(db, doc.id)


# ---------- Extraer datos del oficio externo (no. oficio y fecha) ----------------

def _extraer_texto_archivo(file_path: str, max_pages: int = 3) -> str:
    """Extrae texto plano de un PDF o DOCX.

    Orden de intentos:
      PDF  → pdfplumber → pypdf
      DOCX → python-docx
    Devuelve cadena vacía si ningún método funciona.
    """
    ext = os.path.splitext(file_path)[1].lower()

    # ── DOCX / DOC ────────────────────────────────────────────────────────────
    if ext in (".docx", ".doc"):
        try:
            from docx import Document as DocxDocument
            from docx.oxml.ns import qn as _qn

            doc_obj = DocxDocument(file_path)

            # Recopilar TODO el texto del XML del documento (body + headers + footers).
            # Esto incluye párrafos normales, celdas de tabla Y cuadros de texto
            # flotantes (wps:txbxContent / w:txbxContent) del membrete.
            tokens: list[str] = []

            def _cosechar_texto(xml_element) -> None:
                """Extrae todos los w:t del elemento y sus descendientes."""
                for t_el in xml_element.iter(_qn("w:t")):
                    if t_el.text and t_el.text.strip():
                        tokens.append(t_el.text.strip())

            # Body principal
            _cosechar_texto(doc_obj.element.body)

            # Partes de encabezado y pie de página (contienen el membrete con text boxes)
            from docx.opc.constants import RELATIONSHIP_TYPE as RT
            for rel in doc_obj.part.rels.values():
                if rel.reltype in (
                    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/header",
                    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/footer",
                ):
                    try:
                        _cosechar_texto(rel.target_part._element)
                    except Exception:
                        pass

            return "\n".join(tokens[:500]) if tokens else ""
        except Exception as exc:
            logger.warning("python-docx no pudo leer %s: %s", file_path, exc)
        return ""

    # ── PDF ───────────────────────────────────────────────────────────────────
    if ext != ".pdf":
        return ""

    text = ""
    try:
        import pdfplumber
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages[:max_pages]:
                text += (page.extract_text() or "") + "\n"
        if text.strip():
            return text
    except Exception:
        pass
    try:
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        for page in reader.pages[:max_pages]:
            text += (page.extract_text() or "") + "\n"
    except Exception:
        pass
    return text


# Alias para compatibilidad
def _extraer_texto_pdf(pdf_path: str, max_pages: int = 3) -> str:
    return _extraer_texto_archivo(pdf_path, max_pages)


@router.get(
    "/{doc_id}/extraer-datos-oficio-externo",
    summary="Extraer no. de oficio y fecha del PDF externo subido",
)
async def extraer_datos_oficio_externo(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Lee el PDF del oficio externo guardado, extrae texto y devuelve el número de oficio
    y la fecha detectados mediante expresiones regulares.
    Útil para que el frontend alerte al usuario si difieren de los datos del formulario.
    """
    import re

    doc = await crud_documento.get_with_relations(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    result = {"no_oficio_extraido": None, "fecha_extraida": None}

    if not doc.oficio_externo_url or not os.path.exists(doc.oficio_externo_url):
        return result

    # Intentar extracción del archivo guardado (PDF convertido o DOCX original)
    texto = _extraer_texto_archivo(doc.oficio_externo_url)

    # Si el archivo guardado no dio texto (p.ej. PDF escaneado sin OCR), intentar
    # con el DOCX original en caso de que exista en la misma carpeta
    if not texto.strip():
        upload_dir = os.path.dirname(doc.oficio_externo_url)
        for _ext in (".docx", ".doc"):
            _docx_path = os.path.join(upload_dir, f"oficio_externo_{doc_id}{_ext}")
            if os.path.exists(_docx_path):
                texto = _extraer_texto_archivo(_docx_path)
                break

    if not texto.strip():
        return result

    # ── Detectar número de oficio ─────────────────────────────────────────────
    # Patrones comunes: SFA/DPP/1260/2026, SFA/SF/DPP/SCG/0001/2026, etc.
    # Busca la cadena más larga que empiece por una sigla institucional y
    # termine en /NNN.../20YY (número consecutivo + año).
    folio_match = re.search(
        r'\b(?:SFA|GEM|DPP|SEFOA|SF)(?:/[A-Z0-9]+){1,6}/\d{3,5}/20\d{2}\b',
        texto,
        re.IGNORECASE,
    )
    if folio_match:
        result["no_oficio_extraido"] = folio_match.group(0).upper()

    # ── Detectar fecha en español ─────────────────────────────────────────────
    meses = (
        "enero|febrero|marzo|abril|mayo|junio|julio|agosto|"
        "septiembre|octubre|noviembre|diciembre"
    )
    fecha_match = re.search(
        rf'\b(\d{{1,2}})\s+de\s+({meses})\s+de\s+(20\d{{2}})\b',
        texto,
        re.IGNORECASE,
    )
    if fecha_match:
        result["fecha_extraida"] = fecha_match.group(0)

    return result


@router.delete(
    "/{doc_id}/oficio-externo",
    response_model=DocumentoResponse,
    summary="Eliminar oficio externo subido",
)
async def eliminar_oficio_externo(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    doc = await crud_documento.get_with_relations(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    # Eliminar archivo físico (PDF convertido) si existe
    if doc.oficio_externo_url and os.path.exists(doc.oficio_externo_url):
        try:
            os.remove(doc.oficio_externo_url)
        except OSError:
            pass

    # Eliminar también el DOCX original si quedó en disco (su PDF ya se borró arriba)
    if doc.oficio_externo_nombre:
        upload_dir = os.path.join("uploads", str(doc.cliente_id), "oficios_externos")
        for _ext in (".docx", ".doc"):
            _orig = os.path.join(upload_dir, f"oficio_externo_{doc_id}{_ext}")
            if os.path.exists(_orig):
                try:
                    os.remove(_orig)
                except OSError:
                    pass

    # Limpiar campos
    doc.oficio_externo_url = None
    doc.oficio_externo_nombre = None
    # Si el borrador era un marcador de oficio externo, limpiarlo también
    if doc.borrador_respuesta and doc.borrador_respuesta.startswith("[OFICIO EXTERNO:"):
        doc.borrador_respuesta = None

    db.add(doc)
    await db.commit()
    await db.refresh(doc)
    return await crud_documento.get_with_relations(db, doc.id)


# ---------- Upload de archivo (sin OCR) --------------------------------------

@router.post("/{doc_id}/upload", response_model=DocumentoResponse, summary="Adjuntar archivo")
async def subir_archivo(
    doc_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    if file.content_type not in MIME_PERMITIDOS:
        raise BusinessError(f"Tipo de archivo no permitido: {file.content_type}")

    content = await file.read()
    if len(content) > MAX_SIZE_MB * 1024 * 1024:
        raise BusinessError(f"El archivo supera el limite de {MAX_SIZE_MB} MB.")

    ext = os.path.splitext(file.filename or "doc.pdf")[1]
    filename = f"{uuid.uuid4()}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, "wb") as f:
        f.write(content)

    updated = await crud_documento.actualizar_archivo(
        db,
        db_obj=doc,
        nombre_archivo=file.filename or filename,
        url_storage=filepath,
        mime_type=file.content_type or "application/octet-stream",
    )
    return await crud_documento.get_with_relations(db, updated.id)


# ---------- Cargar documento de referencia para IA --------------------------

@router.post(
    "/{doc_id}/cargar-referencia",
    response_model=DocumentoResponse,
    summary="Subir documento de referencia (tabla, respuesta borrador, info adicional) para enriquecer la generación IA",
)
async def cargar_referencia(
    doc_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    if file.content_type not in MIME_PERMITIDOS:
        raise BusinessError(f"Tipo de archivo no permitido: {file.content_type}")

    content = await file.read()
    if len(content) > MAX_SIZE_MB * 1024 * 1024:
        raise BusinessError(f"El archivo supera el limite de {MAX_SIZE_MB} MB.")

    # Guardar archivo
    ext = os.path.splitext(file.filename or "ref.pdf")[1]
    filename = f"ref_{uuid.uuid4()}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, "wb") as f:
        f.write(content)

    # Extraer texto del documento de referencia
    texto_extraido = ""
    mime = file.content_type or ""

    if mime.startswith("image/") or mime == "image/tiff":
        # Usar Gemini Vision para extraer texto de imagen
        try:
            from app.services.gemini_service import gemini_service
            if gemini_service.available:
                texto_extraido = await gemini_service.extract_text_from_image(content, mime)
            else:
                texto_extraido = "[OCR no disponible — Gemini no configurado]"
        except Exception as e:
            logger.warning(f"Error OCR referencia: {e}")
            texto_extraido = f"[Error al extraer texto de imagen: {e}]"

    elif mime == "application/pdf":
        # Extraer texto de PDF
        try:
            import io
            try:
                import PyPDF2
                reader = PyPDF2.PdfReader(io.BytesIO(content))
                texto_extraido = "\n".join(page.extract_text() or "" for page in reader.pages)
            except ImportError:
                try:
                    import fitz  # PyMuPDF
                    pdf_doc = fitz.open(stream=content, filetype="pdf")
                    texto_extraido = "\n".join(page.get_text() for page in pdf_doc)
                except ImportError:
                    texto_extraido = "[Instalar PyPDF2 o PyMuPDF para extraer texto de PDF]"
        except Exception as e:
            texto_extraido = f"[Error al leer PDF: {e}]"

    elif mime in ("application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                  "application/msword"):
        # Extraer texto de Word — incluyendo tablas
        try:
            import io
            import docx
            doc_word = docx.Document(io.BytesIO(content))
            parts = []
            # Iterar sobre el body en orden (párrafos y tablas intercalados)
            for element in doc_word.element.body:
                if element.tag.endswith('}p'):
                    # Es un párrafo
                    from docx.oxml.ns import qn
                    texts = [node.text or "" for node in element.iter(qn('w:t'))]
                    line = "".join(texts).strip()
                    if line:
                        parts.append(line)
                elif element.tag.endswith('}tbl'):
                    # Es una tabla — extraer preservando estructura
                    from docx.table import Table as DocxTable
                    table = DocxTable(element, doc_word)
                    parts.append("\n[TABLA]")
                    for row in table.rows:
                        cells = [cell.text.strip() for cell in row.cells]
                        parts.append(" | ".join(cells))
                    parts.append("[/TABLA]\n")
            texto_extraido = "\n".join(parts)
        except ImportError:
            texto_extraido = "[Instalar python-docx para extraer texto de Word]"
        except Exception as e:
            texto_extraido = f"[Error al leer DOCX: {e}]"

    elif mime in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                  "application/vnd.ms-excel"):
        # Extraer texto de Excel
        try:
            import io
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            rows = []
            for ws in wb.worksheets:
                rows.append(f"=== Hoja: {ws.title} ===")
                for row in ws.iter_rows(values_only=True):
                    vals = [str(c) if c is not None else "" for c in row]
                    if any(v.strip() for v in vals):
                        rows.append(" | ".join(vals))
            texto_extraido = "\n".join(rows)
        except ImportError:
            texto_extraido = "[Instalar openpyxl para extraer texto de Excel]"
        except Exception as e:
            texto_extraido = f"[Error al leer Excel: {e}]"

    else:
        # Texto plano o formato desconocido — intentar decodificar
        try:
            texto_extraido = content.decode("utf-8", errors="replace")[:50000]
        except Exception:
            texto_extraido = "[Formato no soportado para extracción de texto]"

    # Guardar en el documento
    updated = await crud_documento.update(db, db_obj=doc, obj_in={
        "referencia_archivo_nombre": file.filename or filename,
        "referencia_archivo_url": filepath,
        "contenido_referencia": texto_extraido[:100000],  # limitar a 100k chars
    })
    return await crud_documento.get_with_relations(db, updated.id)


# ---------- Eliminar referencia cargada ------------------------------------

@router.delete(
    "/{doc_id}/referencia",
    response_model=DocumentoResponse,
    summary="Eliminar documento de referencia",
)
async def eliminar_referencia(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))
    updated = await crud_documento.update(db, db_obj=doc, obj_in={
        "referencia_archivo_nombre": None,
        "referencia_archivo_url": None,
        "contenido_referencia": None,
    })
    return await crud_documento.get_with_relations(db, updated.id)


# ---------- Cargar tabla para insertar en DOCX (imagen o Excel) ------------

def _extract_excel_table(content: bytes) -> list[list[str]]:
    """Extrae datos de un archivo Excel como lista de filas (listas de strings)."""
    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        # Ignorar filas completamente vacías
        if all(c is None for c in row):
            continue
        rows.append([str(c) if c is not None else "" for c in row])
    wb.close()
    return rows


ALLOWED_IMAGE_MIMES = ("image/png", "image/jpeg", "image/jpg", "image/webp")
ALLOWED_EXCEL_MIMES = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
)


@router.post(
    "/{doc_id}/tabla-imagen",
    response_model=DocumentoResponse,
    summary="Subir tabla para insertar en el oficio DOCX (imagen PNG/JPG/WEBP o Excel .xlsx)",
)
async def cargar_tabla_imagen(
    doc_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise BusinessError("El archivo no puede pesar más de 10 MB.")

    fname = file.filename or "archivo"
    ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""
    is_excel = file.content_type in ALLOWED_EXCEL_MIMES or ext in ("xlsx", "xls")
    is_image = file.content_type in ALLOWED_IMAGE_MIMES or ext in ("png", "jpg", "jpeg", "webp")

    if not is_excel and not is_image:
        raise BusinessError(
            f"Formato no soportado ({file.content_type}). "
            "Usa imágenes (PNG, JPG, WEBP) o archivos Excel (.xlsx)."
        )

    update_data: dict = {}

    if is_excel:
        # Extraer datos del Excel como tabla estructurada
        try:
            table_data = _extract_excel_table(content)
        except Exception as exc:
            raise BusinessError(f"No se pudo leer el archivo Excel: {exc}")
        if not table_data:
            raise BusinessError("El archivo Excel está vacío o no tiene datos.")
        update_data["tabla_datos_json"] = table_data
        update_data["tabla_imagen_url"] = None  # limpiar imagen si había
        update_data["tabla_imagen_nombre"] = fname
    else:
        # Guardar imagen
        os.makedirs("uploads/tablas", exist_ok=True)
        img_filename = f"tabla_{doc_id}.{ext or 'png'}"
        filepath = f"uploads/tablas/{img_filename}"
        with open(filepath, "wb") as f:
            f.write(content)
        update_data["tabla_imagen_url"] = filepath
        update_data["tabla_imagen_nombre"] = fname
        update_data["tabla_datos_json"] = None  # limpiar datos Excel si había

    updated = await crud_documento.update(db, db_obj=doc, obj_in=update_data)
    return await crud_documento.get_with_relations(db, updated.id)


@router.delete(
    "/{doc_id}/tabla-imagen",
    response_model=DocumentoResponse,
    summary="Eliminar tabla (imagen o datos Excel)",
)
async def eliminar_tabla_imagen(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))
    updated = await crud_documento.update(db, db_obj=doc, obj_in={
        "tabla_imagen_url": None,
        "tabla_imagen_nombre": None,
        "tabla_datos_json": None,
    })
    return await crud_documento.get_with_relations(db, updated.id)


# ---------- Generar oficio estructurado (4 secciones) ----------------------

@router.post(
    "/{doc_id}/generar-oficio-estructurado",
    response_model=OficioEstructuradoResponse,
    summary="Generar borrador con estructura juridica de 4 secciones",
)
async def generar_oficio_estructurado(
    doc_id: str,
    body: Optional[dict] = None,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Genera un borrador de oficio de respuesta con estructura juridica obligatoria:
    1. Fundamento competencial
    2. Referencia del oficio
    3. Objeto del oficio
    4. Cierre institucional

    Inyecta contexto normativo (RISFA, Ley Disciplina, LGCG, Decreto) al prompt.
    """
    doc = await crud_documento.get_with_relations(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    if doc.flujo != "recibido":
        raise BusinessError("Solo los oficios recibidos tienen flujo de respuesta.")

    instrucciones = ""
    if body and isinstance(body, dict):
        instrucciones = body.get("instrucciones", "")

    # Construir contexto normativo
    from app.services.normativa_context_service import normativa_context_service
    contexto_normativo = await normativa_context_service.build_context_for_oficio(
        db,
        regla_turno_codigo=doc.regla_turno_codigo,
        area_codigo=doc.area_turno,
        asunto=doc.asunto or "",
    )

    area_nombre = doc.area_turno_nombre or doc.sugerencia_area_nombre or "Direccion de Programacion y Presupuesto"
    fundamento = doc.sugerencia_fundamento or "Arts. 18 y 19 del Reglamento Interior de la SFA"
    resumen_ocr = (doc.datos_extraidos_ia or {}).get("cuerpo_resumen", "")

    secciones = await correspondencia_service.generar_oficio_estructurado(
        numero_oficio_origen=doc.numero_oficio_origen or "---",
        fecha_recibido=doc.fecha_recibido or doc.fecha_documento or "---",
        remitente_nombre=doc.remitente_nombre or "---",
        remitente_cargo=doc.remitente_cargo or "---",
        remitente_dependencia=doc.remitente_dependencia or "---",
        asunto=doc.asunto,
        cuerpo_resumen=resumen_ocr,
        area_nombre=area_nombre,
        fundamento_legal=fundamento,
        contexto_normativo=contexto_normativo,
        instrucciones=instrucciones,
    )

    # Concatenar secciones para el borrador completo
    borrador_completo = "\n\n".join(
        s for s in [
            secciones.get("fundamento", ""),
            secciones.get("referencia", ""),
            secciones.get("objeto", ""),
            secciones.get("cierre", ""),
        ] if s and s.strip()
    )

    # Guardar borrador en BD
    await crud_documento.guardar_borrador(db, db_obj=doc, borrador=borrador_completo)

    return OficioEstructuradoResponse(
        secciones=secciones,
        borrador_completo=borrador_completo,
        message="Oficio estructurado generado con 4 secciones juridicas.",
    )


# ---------- Descargar oficio DOCX -------------------------------------------

@router.post(
    "/{doc_id}/descargar-oficio",
    summary="Generar y descargar oficio en formato DOCX",
)
async def descargar_oficio(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Genera un documento DOCX con formato institucional y lo retorna para descarga.
    Roles con permiso de edición de respuesta pueden descargar en DOCX para editar en Word."""
    _ROLES_DOCX = {"superadmin", "admin_cliente", "secretaria", "analista",
                   "subdirector", "jefe_depto", "asesor"}
    if current_user.rol not in _ROLES_DOCX:
        raise ForbiddenError("No tiene permisos para descargar en formato Word.")
    doc = await crud_documento.get_with_relations(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    if not doc.borrador_respuesta:
        raise BusinessError("No hay borrador de respuesta. Genere uno primero.")

    from app.services.oficio_generator_service import oficio_generator
    from app.services.correspondencia_service import CorrespondenciaService
    from datetime import date

    # Parsear secciones del borrador
    secciones = CorrespondenciaService._parse_secciones(doc.borrador_respuesta)

    # Formatear fecha — usar fecha guardada si existe, si no usar hoy
    from app.services.oficio_generator_service import MESES_ES
    if doc.fecha_respuesta:
        fecha_str = doc.fecha_respuesta
    else:
        hoy = date.today()
        fecha_str = f"{hoy.day} de {MESES_ES[hoy.month - 1]} de {hoy.year}"

    # --- Preparar datos de sello digital si está firmado ---
    sello_data = None
    if doc.firmado_digitalmente and doc.firma_metadata:
        import json as _json
        meta = doc.firma_metadata
        if isinstance(meta, str):
            try:
                meta = _json.loads(meta)
            except Exception:
                meta = {}

        # Leer QR PNG desde archivo
        qr_png_bytes = b""
        qr_path = meta.get("qr_url", "")
        if qr_path:
            from pathlib import Path as _Path
            qr_file = _Path(qr_path)
            if qr_file.exists():
                qr_png_bytes = qr_file.read_bytes()
            else:
                # Intentar ruta alternativa
                qr_alt = _Path(f"uploads/qr_firmas/qr_{doc_id}.png")
                if qr_alt.exists():
                    qr_png_bytes = qr_alt.read_bytes()

        sello_data = {
            "qr_png_bytes": qr_png_bytes,
            "nombre_firmante": meta.get("nombre_firmante", ""),
            "cargo_firmante": meta.get("cargo", "Director de Programación y Presupuesto"),
            "rfc_firmante": meta.get("rfc_firmante", ""),
            "serial_certificado": meta.get("serial_certificado", ""),
            "fecha_firma": meta.get("fecha_firma", ""),
            "correo_firmante": meta.get("firmado_por_usuario", ""),
            "folio_firma": doc.folio_respuesta or "",
        }

    # --- Determinar destinatario según flujo ---
    if doc.flujo == "emitido":
        if doc.destinatario_nombre:
            dest_nombre = doc.destinatario_nombre
            dest_cargo = doc.destinatario_cargo or ""
            dest_dep = doc.dependencia_destino or ""
        else:
            # Legacy: solo dependencia_destino → ponerlo como nombre, sin duplicar
            dest_nombre = doc.dependencia_destino or "---"
            dest_cargo = ""
            dest_dep = ""
    else:
        # Para recibidos: usar destinatario_nombre/cargo/dependencia_destino si el usuario
        # los personalizó (respuesta a alguien distinto al remitente original).
        # Fallback al remitente del documento original si no se personalizó.
        dest_nombre = doc.destinatario_nombre or doc.remitente_nombre or "---"
        dest_cargo  = doc.destinatario_cargo  or doc.remitente_cargo  or "---"
        dest_dep    = doc.dependencia_destino or doc.remitente_dependencia or "---"

    # --- Determinar copias estándar según área ---
    area_codigo = doc.area_turno or "DIR"
    plantilla = detectar_plantilla(doc.asunto or "", "")
    if plantilla:
        copias_doc = obtener_copias(plantilla.get("copias", "presupuestales"))
    else:
        # Áreas de subdirección → presupuestales; DIR → administrativas
        copias_doc = obtener_copias(
            "administrativas" if area_codigo == "DIR" else "presupuestales"
        )

    # Tabla: imagen o datos Excel extraídos
    tabla_img = doc.tabla_imagen_url if hasattr(doc, 'tabla_imagen_url') and doc.tabla_imagen_url and os.path.exists(doc.tabla_imagen_url) else None
    tabla_datos = doc.tabla_datos_json if hasattr(doc, 'tabla_datos_json') else None

    docx_bytes = oficio_generator.generar_oficio_respuesta(
        folio_respuesta=doc.folio_respuesta or "SFA/SF/DPP/____/2026",
        fecha_respuesta=fecha_str,
        destinatario_nombre=dest_nombre,
        destinatario_cargo=dest_cargo,
        destinatario_dependencia=dest_dep,
        seccion_fundamento=secciones.get("fundamento", ""),
        seccion_referencia=secciones.get("referencia", ""),
        seccion_objeto=secciones.get("objeto", ""),
        seccion_cierre=secciones.get("cierre", ""),
        firmante_nombre="Mtro. Marco Antonio Flores Mejía",
        firmante_cargo="Director de Programación y Presupuesto",
        referencia_elaboro=doc.referencia_elaboro,
        referencia_reviso=doc.referencia_reviso,
        copias=copias_doc,
        incluir_firma_visual=bool(doc.firmado_digitalmente),
        sello_digital_data=sello_data,
        asunto=doc.asunto,
        tabla_imagen_path=tabla_img,
        tabla_datos_json=tabla_datos,
    )

    import io
    import urllib.parse

    # Construir nombre con nomenclatura oficial: OF. RESP. {consecutivo}-{año}
    # El folio tiene formato tipo "SFA/SF/DPP/1260/2026"; tomamos las dos últimas partes.
    folio_raw = (doc.folio_respuesta or "").strip()
    try:
        partes = [p.strip() for p in folio_raw.split("/") if p.strip()]
        if len(partes) >= 2:
            consecutivo = partes[-2]
            anio        = partes[-1]
            nombre_archivo = f"OF. RESP. {consecutivo}-{anio}.docx"
        elif folio_raw:
            nombre_archivo = f"OF. RESP. {folio_raw}.docx"
        else:
            nombre_archivo = "OF. RESP. oficio.docx"
    except Exception:
        nombre_archivo = "OF. RESP. oficio.docx"

    # RFC 5987 para que el nombre con espacios y puntos se descargue correctamente
    encoded_name = urllib.parse.quote(nombre_archivo, safe="")
    return StreamingResponse(
        io.BytesIO(docx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{nombre_archivo}"; '
                f"filename*=UTF-8''{encoded_name}"
            )
        },
    )


# ---------- Señal de módulo externo (certificaciones, etc.) -------------------

@router.post(
    "/{doc_id}/modulo-externo",
    response_model=DocumentoResponse,
    summary="Recibir señal de módulo externo (certificaciones) para actualizar estado",
)
async def señal_modulo_externo(
    doc_id: str,
    estado: str = Query(..., description="Estado a establecer (ej: 'firmado')"),
    referencia: str = Query("", description="Referencia del módulo externo (ej: folio certificación)"),
    modulo: str = Query("certificaciones", description="Nombre del módulo origen"),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Permite que otro módulo (certificaciones presupuestales, minutas)
    actualice el estado de un documento en gestión documental.
    Ejemplo: certificación atendida → documento pasa a 'firmado'.
    """
    if current_user.rol not in ("admin_cliente", "superadmin"):
        raise ForbiddenError("Solo admin o superadmin pueden enviar señales entre módulos.")

    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    updated = await crud_documento.update(db, db_obj=doc, obj_in={
        "estado": estado,
        "modulo_externo_estado": f"atendido_{modulo}",
        "modulo_externo_ref": referencia,
    })
    return await crud_documento.get_with_relations(db, updated.id)


# ---------- Devolver documento -----------------------------------------------

@router.post(
    "/{doc_id}/devolver",
    response_model=DevolucionResponse,
    summary="Devolver documento al área responsable",
)
async def devolver_documento(
    doc_id: str,
    data: DevolucionInput,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Director devuelve un documento al área responsable para correcciones.
    Solo admin_cliente o superadmin.
    Transición: en_atencion/respondido → devuelto (recibidos)
                en_revision → borrador (emitidos)
    """
    if current_user.rol not in ("admin_cliente", "superadmin", "subdirector"):
        raise ForbiddenError("Solo el Director, Subdirector o superadmin pueden devolver documentos.")

    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    # Emitidos en revisión: devolver a borrador con motivo
    if doc.flujo == "emitido" and doc.estado == "en_revision":
        if not doc.borrador_respuesta and not doc.url_storage:
            raise BusinessError(
                "No hay contenido que devolver. El documento emitido no tiene "
                "borrador ni archivo adjunto."
            )
        doc = await crud_documento.devolver_documento(
            db,
            db_obj=doc,
            observaciones=data.observaciones,
            devuelto_por_id=str(current_user.id),
            estado_destino="borrador",
        )
        historial = await crud_documento.get_historial(db, doc.id)
        entry = historial[0] if historial else None
        return DevolucionResponse(
            documento_id=doc.id,
            estado="borrador",
            historial_entry=HistorialItemResponse(
                id=entry.id if entry else "",
                tipo_accion=entry.tipo_accion if entry else "devolucion",
                estado_anterior=entry.estado_anterior if entry else None,
                estado_nuevo=entry.estado_nuevo if entry else None,
                observaciones=entry.observaciones if entry else data.observaciones,
                version=entry.version if entry else 1,
                timestamp=entry.timestamp if entry else "",
                usuario_nombre=(
                    entry.usuario.nombre_completo if entry and entry.usuario else None
                ),
            ) if entry else None,
            message="Documento emitido devuelto a borrador.",
        )

    if doc.estado not in ("en_atencion", "respondido"):
        raise BusinessError(
            f"Solo documentos 'en_atencion', 'respondido' o 'en_revision' (emitidos) pueden devolverse. Estado actual: {doc.estado}"
        )
    if not doc.borrador_respuesta:
        raise BusinessError("No hay borrador que devolver. El documento no tiene respuesta generada.")

    doc = await crud_documento.devolver_documento(
        db,
        db_obj=doc,
        observaciones=data.observaciones,
        devuelto_por_id=str(current_user.id),
    )

    # Obtener la entrada de historial recién creada
    historial = await crud_documento.get_historial(db, doc.id)
    entry = historial[0] if historial else None

    return DevolucionResponse(
        documento_id=doc.id,
        estado="devuelto",
        historial_entry=HistorialItemResponse(
            id=entry.id,
            tipo_accion=entry.tipo_accion,
            estado_anterior=entry.estado_anterior,
            estado_nuevo=entry.estado_nuevo,
            observaciones=entry.observaciones,
            version=entry.version,
            timestamp=entry.timestamp,
            usuario_nombre=(
                entry.usuario.nombre_completo if entry.usuario else None
            ),
        ) if entry else None,
        message="Documento devuelto al área responsable.",
    )


# ---------- Reenviar documento corregido ------------------------------------

@router.post(
    "/{doc_id}/reenviar",
    response_model=DocumentoResponse,
    summary="Reenviar documento corregido para revisión",
)
async def reenviar_documento(
    doc_id: str,
    data: ReenvioInput,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Área responsable reenvía documento corregido al Director.
    Transición: devuelto → en_atencion. Incrementa versión.
    """
    if current_user.rol == "consulta":
        raise ForbiddenError("No tienes permisos para reenviar documentos.")

    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    if doc.estado != "devuelto":
        raise BusinessError(
            f"Solo documentos 'devuelto' pueden reenviarse. Estado actual: {doc.estado}"
        )

    doc = await crud_documento.reenviar_documento(
        db,
        db_obj=doc,
        comentario=data.comentario or "",
        reenviado_por_id=str(current_user.id),
    )
    return await crud_documento.get_with_relations(db, doc.id)


# ---------- Historial de un documento ---------------------------------------

@router.get(
    "/{doc_id}/historial",
    response_model=List[HistorialItemResponse],
    summary="Obtener historial de acciones del documento",
)
async def obtener_historial(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Timeline de acciones: devoluciones, reenvíos, firmas."""
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    historial = await crud_documento.get_historial(db, doc_id)
    return [
        HistorialItemResponse(
            id=h.id,
            tipo_accion=h.tipo_accion,
            estado_anterior=h.estado_anterior,
            estado_nuevo=h.estado_nuevo,
            observaciones=h.observaciones,
            version=h.version,
            timestamp=h.timestamp,
            usuario_nombre=(
                h.usuario.nombre_completo if h.usuario else None
            ),
        )
        for h in historial
    ]


# ---------- Firmar documento ------------------------------------------------

@router.post(
    "/{doc_id}/firmar",
    response_model=FirmaResponse,
    summary="Firmar documento con e.firma",
)
async def firmar_documento(
    doc_id: str,
    request: Request,
    password: str = Form("", description="Contraseña de la clave privada"),
    cer_file: Optional[UploadFile] = File(None, description="Archivo .cer (opcional si tiene certificado registrado)"),
    key_file: Optional[UploadFile] = File(None, description="Archivo .key (opcional si tiene certificado registrado)"),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Firma un documento con e.firma electrónica avanzada.

    Si el usuario tiene certificado registrado en la bóveda, solo necesita
    la contraseña. Si no, puede subir .cer + .key + password.

    Genera firma RSA real, QR de verificación y registra en bitácora.
    """
    if current_user.rol != "admin_cliente":
        raise ForbiddenError("Solo el Director puede firmar documentos.")

    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    if not doc.borrador_respuesta:
        raise BusinessError("No hay borrador de respuesta. Genere uno antes de firmar.")

    if doc.estado == "devuelto":
        raise BusinessError(
            "Este documento fue devuelto y requiere correcciones antes de firmarse. "
            "Corrija el borrador y reenvíe antes de intentar firmar."
        )
    # Estados válidos para firma: cualquier estado con borrador listo
    estados_firmables = ("en_atencion", "respondido", "borrador", "en_revision", "turnado")
    if doc.estado not in estados_firmables:
        raise BusinessError(
            f"El documento debe estar en un estado válido para firmarse. Estado actual: {doc.estado}"
        )

    from app.services.firma_electronica_service import firma_electronica_service
    from app.services.boveda_certificados_service import boveda_certificados_service

    usuario_id = str(current_user.id)
    key_bytes = b""
    cer_bytes = b""
    cert_info = None

    # Intentar obtener certificado de la bóveda
    cert_record = await boveda_certificados_service.obtener_certificado(db, usuario_id)

    if cert_record and password:
        # Usar certificado de la bóveda
        try:
            key_bytes, cer_bytes, cert_record = await boveda_certificados_service.descifrar_clave_privada(
                db, usuario_id=usuario_id, password=password,
                ip_origen=request.client.host if request.client else "",
            )
            cert_info = {
                "serial": cert_record.numero_serie,
                "rfc": cert_record.rfc,
                "nombre": cert_record.nombre_titular,
                "valido_desde": cert_record.valido_desde.isoformat() if cert_record.valido_desde else "",
                "valido_hasta": cert_record.valido_hasta.isoformat() if cert_record.valido_hasta else "",
            }
        except ValueError as e:
            raise BusinessError(str(e))
    elif cer_file and key_file and password:
        # Fallback: subir archivos directamente
        cer_bytes = await cer_file.read()
        key_bytes = await key_file.read()
        cert_result = firma_electronica_service.validar_certificado(cer_bytes, key_bytes, password)
        if not cert_result.get("valido"):
            raise BusinessError("Certificado inválido: " + cert_result.get("message", ""))
        cert_info = {
            "serial": cert_result["serial"],
            "rfc": cert_result["rfc"],
            "nombre": cert_result["nombre"],
            "valido_desde": cert_result.get("valido_desde", ""),
            "valido_hasta": cert_result.get("valido_hasta", ""),
        }
    else:
        raise BusinessError(
            "Debe proporcionar la contraseña de su e.firma para firmar. "
            "Si no tiene certificado registrado, suba los archivos .cer y .key."
        )

    # Auto-generar folio si no tiene
    if not doc.folio_respuesta:
        from datetime import datetime as dt_
        from sqlalchemy import text as sql_text

        anio_ = dt_.now().year
        area_ = (doc.area_turno or "DIR").upper()
        prefijo_ = PREFIJOS_FOLIO.get(area_, "SFA/SF/DPP")
        pattern_ = f"{prefijo_}/%/{anio_}"

        result_ = await db.execute(
            sql_text(
                "SELECT folio_respuesta FROM documentos_oficiales "
                "WHERE folio_respuesta LIKE :pattern "
                "ORDER BY folio_respuesta DESC LIMIT 1"
            ),
            {"pattern": pattern_},
        )
        row_ = result_.first()
        next_num_ = 1
        if row_ and row_[0]:
            try:
                parts_ = row_[0].split("/")
                next_num_ = int(parts_[-2]) + 1
            except (ValueError, IndexError):
                pass

        folio_auto = f"{prefijo_}/{str(next_num_).zfill(4)}/{anio_}"
        doc.folio_respuesta = folio_auto
        await db.flush()

    # Firmar con criptografía real
    firma_result = firma_electronica_service.firmar_documento(
        contenido_borrador=doc.borrador_respuesta,
        serial_certificado=cert_info["serial"],
        rfc_firmante=cert_info["rfc"],
        nombre_firmante=cert_info["nombre"],
        folio=doc.folio_respuesta or "SIN-FOLIO",
        certificado_valido_desde=cert_info.get("valido_desde", ""),
        certificado_valido_hasta=cert_info.get("valido_hasta", ""),
        key_bytes=key_bytes,
        password=password,
    )
    firma_result["firmado_por_usuario"] = current_user.email
    firma_result["cargo"] = "Director de Programación y Presupuesto"

    # Generar QR de verificación
    qr_bytes, qr_json = firma_electronica_service.generar_qr(
        hash_documento=firma_result["hash_documento"],
        fecha_firma=firma_result["fecha_firma"],
        serial_certificado=cert_info["serial"],
        valido_desde=cert_info.get("valido_desde", ""),
        valido_hasta=cert_info.get("valido_hasta", ""),
        folio=doc.folio_respuesta or "SIN-FOLIO",
        documento_id=doc.id,
    )
    qr_url = firma_electronica_service.guardar_qr(doc.id, qr_bytes)
    firma_result["qr_url"] = qr_url
    firma_result["qr_data"] = qr_json

    # Actualizar en BD — recibidos pasan a "firmado", emitidos a "vigente"
    estado_final = "firmado" if doc.flujo == "recibido" else "vigente"
    update_data = DocumentoUpdate(
        firmado_digitalmente=True,
        firma_metadata=firma_result,
        estado=estado_final,
    )
    await crud_documento.actualizar_documento(db, db_obj=doc, obj_in=update_data)

    # Registrar en historial
    await crud_documento.registrar_firma_historial(
        db,
        documento_id=doc.id,
        usuario_id=usuario_id,
        version=doc.version or 1,
        estado_anterior=doc.estado,
        estado_nuevo=estado_final,
    )

    # Actualizar contador de firmas en bóveda
    if cert_record:
        await boveda_certificados_service.incrementar_firmas(db, usuario_id)

    # Registrar en bitácora de firma
    await boveda_certificados_service._registrar_bitacora(
        db, usuario_id=usuario_id, accion="firma_individual",
        exitoso=True, documento_id=doc.id,
        rfc=cert_info["rfc"], serial=cert_info["serial"],
        hash_doc=firma_result["hash_documento"],
        ip_origen=request.client.host if request.client else "",
        detalle=f"Documento firmado: {doc.asunto[:50]}",
    )
    await db.commit()

    return FirmaResponse(
        firmado_digitalmente=True,
        firma_metadata=firma_result,
        message="Documento firmado exitosamente con e.firma. QR de verificación generado.",
    )


# ---------- Constancia de firma PDF -------------------------------------------

@router.get(
    "/{doc_id}/constancia-firma",
    summary="Descargar constancia de firma electrónica (PDF)",
)
async def descargar_constancia_firma(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Genera y descarga un PDF con la constancia de firma electrónica.
    Incluye datos criptográficos, QR de verificación y datos del documento.
    """
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    if not doc.firmado_digitalmente or not doc.firma_metadata:
        raise BusinessError("El documento no ha sido firmado digitalmente.")

    from app.services.constancia_firma_service import constancia_firma_service

    # Extraer firma_metadata (puede ser dict o JSON string)
    meta = doc.firma_metadata
    if isinstance(meta, str):
        import json
        meta = json.loads(meta)

    pdf_bytes = constancia_firma_service.generar_constancia_pdf(
        documento_id=doc.id,
        asunto=doc.asunto,
        folio_respuesta=doc.folio_respuesta or "",
        numero_oficio_origen=doc.numero_oficio_origen or "",
        remitente_nombre=doc.remitente_nombre or "",
        remitente_dependencia=doc.remitente_dependencia or "",
        area_turno_nombre=doc.area_turno_nombre or "",
        firma_metadata=meta,
        version=doc.version or 1,
    )

    import io
    filename = f"constancia_firma_{doc.folio_respuesta or doc.id}.pdf"

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


# ---------- Oficio en PDF (vista previa / descarga) --------------------------

@router.get(
    "/{doc_id}/descargar-oficio-pdf",
    summary="Generar y descargar oficio en formato PDF",
)
async def descargar_oficio_pdf(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """
    Genera un PDF del oficio de respuesta con formato institucional.
    Incluye sello digital / QR si el documento está firmado.
    Ideal para vista previa embebida o descarga directa.
    """
    doc = await crud_documento.get_with_relations(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    # ── Si hay un oficio externo ya subido, servirlo directamente sin regenerar ──
    if doc.oficio_externo_url and os.path.exists(doc.oficio_externo_url):
        import io as _io
        with open(doc.oficio_externo_url, "rb") as _f:
            _pdf_bytes = _f.read()
        # Determinar media type: si es PDF sirve como tal; si es DOCX usar octet-stream
        _ext = os.path.splitext(doc.oficio_externo_url)[1].lower()
        _media = "application/pdf" if _ext == ".pdf" else "application/octet-stream"
        _fname = doc.oficio_externo_nombre or os.path.basename(doc.oficio_externo_url)
        return StreamingResponse(
            _io.BytesIO(_pdf_bytes),
            media_type=_media,
            headers={"Content-Disposition": f'inline; filename="{_fname}"'},
        )

    if not doc.borrador_respuesta:
        raise BusinessError("No hay borrador de respuesta. Genere uno primero.")

    from app.services.oficio_pdf_service import oficio_pdf_service
    from app.services.correspondencia_service import CorrespondenciaService
    from app.services.oficio_generator_service import MESES_ES
    from datetime import date
    import json as _json

    # Parsear secciones
    secciones = CorrespondenciaService._parse_secciones(doc.borrador_respuesta)

    # Fecha — usar fecha guardada si existe, si no usar hoy
    if doc.fecha_respuesta:
        fecha_str = doc.fecha_respuesta
    else:
        hoy = date.today()
        fecha_str = f"{hoy.day} de {MESES_ES[hoy.month - 1]} de {hoy.year}"

    # Datos de sello digital si está firmado
    sello_data = None
    if doc.firmado_digitalmente and doc.firma_metadata:
        meta = doc.firma_metadata
        if isinstance(meta, str):
            try:
                meta = _json.loads(meta)
            except Exception:
                meta = {}

        qr_png_bytes = b""
        qr_path = meta.get("qr_url", "")
        if qr_path:
            from pathlib import Path as _Path
            qr_file = _Path(qr_path)
            if qr_file.exists():
                qr_png_bytes = qr_file.read_bytes()
            else:
                qr_alt = _Path(f"uploads/qr_firmas/qr_{doc_id}.png")
                if qr_alt.exists():
                    qr_png_bytes = qr_alt.read_bytes()

        sello_data = {
            "qr_png_bytes": qr_png_bytes,
            "nombre_firmante": meta.get("nombre_firmante", ""),
            "cargo_firmante": meta.get("cargo", "Director de Programación y Presupuesto"),
            "rfc_firmante": meta.get("rfc_firmante", ""),
            "serial_certificado": meta.get("serial_certificado", ""),
            "fecha_firma": meta.get("fecha_firma", ""),
            "correo_firmante": meta.get("firmado_por_usuario", ""),
            "folio_firma": doc.folio_respuesta or "",
        }

    # --- Determinar destinatario según flujo ---
    if doc.flujo == "emitido":
        if doc.destinatario_nombre:
            # Tiene nombre separado → usar nombre/cargo/dependencia
            pdf_dest_nombre = doc.destinatario_nombre
            pdf_dest_cargo = doc.destinatario_cargo or ""
            pdf_dest_dep = doc.dependencia_destino or ""
        else:
            # Solo tiene dependencia_destino (legacy) → ponerlo como nombre, sin duplicar
            pdf_dest_nombre = doc.dependencia_destino or "---"
            pdf_dest_cargo = ""
            pdf_dest_dep = ""
    else:
        # Usar destinatario_* si el usuario personalizó la respuesta hacia otra persona.
        # Fallback al remitente original del documento recibido.
        pdf_dest_nombre = doc.destinatario_nombre or doc.remitente_nombre or "---"
        pdf_dest_cargo  = doc.destinatario_cargo  or doc.remitente_cargo  or "---"
        pdf_dest_dep    = doc.dependencia_destino or doc.remitente_dependencia or "---"

    # --- Determinar copias estándar según área ---
    pdf_area = doc.area_turno or "DIR"
    pdf_plantilla = detectar_plantilla(doc.asunto or "", "")
    if pdf_plantilla:
        pdf_copias = obtener_copias(pdf_plantilla.get("copias", "presupuestales"))
    else:
        pdf_copias = obtener_copias(
            "administrativas" if pdf_area == "DIR" else "presupuestales"
        )

    # Tabla: imagen o datos Excel extraídos (igual que DOCX)
    pdf_tabla_img = doc.tabla_imagen_url if hasattr(doc, 'tabla_imagen_url') and doc.tabla_imagen_url and os.path.exists(doc.tabla_imagen_url) else None
    pdf_tabla_datos = doc.tabla_datos_json if hasattr(doc, 'tabla_datos_json') else None

    pdf_bytes = oficio_pdf_service.generar_oficio_pdf(
        folio_respuesta=doc.folio_respuesta or "SFA/SF/DPP/____/2026",
        fecha_respuesta=fecha_str,
        destinatario_nombre=pdf_dest_nombre,
        destinatario_cargo=pdf_dest_cargo,
        destinatario_dependencia=pdf_dest_dep,
        seccion_fundamento=secciones.get("fundamento", ""),
        seccion_referencia=secciones.get("referencia", ""),
        seccion_objeto=secciones.get("objeto", ""),
        seccion_cierre=secciones.get("cierre", ""),
        firmante_nombre="Mtro. Marco Antonio Flores Mejía",
        firmante_cargo="Director de Programación y Presupuesto",
        referencia_elaboro=doc.referencia_elaboro,
        referencia_reviso=doc.referencia_reviso,
        copias=pdf_copias,
        incluir_firma_visual=bool(doc.firmado_digitalmente),
        sello_digital_data=sello_data,
        asunto=doc.asunto,
        tabla_imagen_path=pdf_tabla_img,
        tabla_datos_json=pdf_tabla_datos,
    )

    import io
    folio_safe = (doc.folio_respuesta or "oficio").replace("/", "_")
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="oficio_{folio_safe}.pdf"',
        },
    )


# ---------- Ver archivo original (turnado/escaneado) --------------------------

@router.get("/{doc_id}/archivo-original", summary="Obtener archivo original del documento")
async def obtener_archivo_original(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Devuelve el archivo original (PDF/imagen) del documento."""
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))

    if not doc.url_storage:
        raise BusinessError("Este documento no tiene archivo adjunto.")

    import io
    from pathlib import Path as _Path
    filepath = _Path(doc.url_storage)
    if not filepath.exists():
        raise BusinessError("El archivo no se encontró en el servidor.")

    content = filepath.read_bytes()
    mime = doc.mime_type or "application/pdf"
    filename = doc.nombre_archivo or f"documento_{doc_id}"

    return StreamingResponse(
        io.BytesIO(content),
        media_type=mime,
        headers={
            "Content-Disposition": f'inline; filename="{filename}"',
        },
    )


# ---------- Eliminar ---------------------------------------------------------

@router.delete("/{doc_id}", response_model=MessageResponse, summary="Eliminar documento")
async def eliminar_documento(
    doc_id: str,
    confirmar: str = Query("no", description="Debe enviar 'si' para confirmar eliminación"),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    # Solo secretaría y admin pueden eliminar
    if current_user.rol not in ("secretaria", "admin_cliente", "superadmin"):
        raise ForbiddenError("Solo Secretaría y Administrador pueden eliminar documentos.")
    if confirmar != "si":
        raise BusinessError("Debe confirmar la eliminación enviando confirmar='si'.")
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))
    await crud_documento.delete(db, id=doc_id)
    return {"message": "Documento eliminado correctamente.", "success": True}


# ---------- Acuse de recibido (secretaria sube escaneo con sello) -----------

@router.post(
    "/{doc_id}/acuse-recibido",
    response_model=DocumentoResponse,
    summary="Subir acuse de recibido (escaneo con sello)",
)
async def subir_acuse_recibido(
    doc_id: str,
    fecha_acuse: str = Form(""),
    archivo: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    """Secretaria sube el escaneo del oficio con sello de acuse de la dependencia destino.
    La fecha del acuse se extrae automáticamente del sello vía IA."""
    if current_user.rol not in ("secretaria", "superadmin"):
        raise ForbiddenError("Solo Secretaría puede registrar acuses de recibido.")
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))
    if not doc.firmado_digitalmente:
        raise BusinessError("Solo documentos firmados pueden recibir acuse.")

    upload_dir = os.path.join("uploads", "acuses")
    os.makedirs(upload_dir, exist_ok=True)
    ext = os.path.splitext(archivo.filename or "acuse.pdf")[1]
    filename = f"acuse_{doc_id}{ext}"
    filepath = os.path.join(upload_dir, filename)
    content = await archivo.read()
    with open(filepath, "wb") as f:
        f.write(content)

    # ── Extraer fecha del sello de acuse vía Gemini Vision ──
    fecha_extraida = fecha_acuse or None
    try:
        import google.generativeai as genai
        api_key = os.environ.get("GEMINI_API_KEY", "")
        if api_key and api_key != "placeholder":
            genai.configure(api_key=api_key)
            mime = "image/jpeg" if ext.lower() in (".jpg", ".jpeg") else "image/png" if ext.lower() == ".png" else "application/pdf"
            model = genai.GenerativeModel("gemini-2.0-flash")
            prompt = (
                "Analiza este documento escaneado. Busca el SELLO DE ACUSE DE RECIBIDO "
                "(generalmente un sello oficial con fecha de recepción). "
                "Extrae SOLO la fecha del sello de acuse en formato 'DD de mes de YYYY'. "
                "Si encuentras la fecha, responde SOLO con la fecha. "
                "Si no encuentras sello de acuse, responde 'SIN_FECHA'."
            )
            response = model.generate_content([
                prompt,
                {"mime_type": mime, "data": content}
            ])
            fecha_text = response.text.strip()
            if fecha_text and fecha_text != "SIN_FECHA" and len(fecha_text) < 60:
                fecha_extraida = fecha_text
    except Exception:
        pass  # Si falla la IA, continúa sin fecha

    from datetime import datetime as dt
    update_data = {
        "acuse_recibido_url": filepath,
        "acuse_recibido_nombre": archivo.filename,
        "acuse_recibido_fecha": fecha_extraida,
        "acuse_registrado_en": dt.now(),
        "acuse_registrado_por_id": str(current_user.id),
        "despachado": True,
        "despachado_por_id": str(current_user.id),
        "despachado_en": dt.now(),
    }
    updated = await crud_documento.update(db, db_obj=doc, obj_in=update_data)
    return await crud_documento.get_with_relations(db, updated.id)


@router.delete(
    "/{doc_id}/acuse-recibido",
    response_model=DocumentoResponse,
    summary="Eliminar acuse de recibido",
)
async def eliminar_acuse_recibido(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    if current_user.rol not in ("secretaria", "superadmin"):
        raise ForbiddenError("Solo Secretaría puede eliminar acuses.")
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))
    if doc.acuse_recibido_url and os.path.exists(doc.acuse_recibido_url):
        os.remove(doc.acuse_recibido_url)
    updated = await crud_documento.update(db, db_obj=doc, obj_in={
        "acuse_recibido_url": None, "acuse_recibido_nombre": None,
        "acuse_recibido_fecha": None, "acuse_registrado_en": None,
        "acuse_registrado_por_id": None,
    })
    return await crud_documento.get_with_relations(db, updated.id)


@router.get("/{doc_id}/acuse-recibido/archivo", summary="Descargar acuse de recibido")
async def descargar_acuse_recibido(
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
):
    doc = await crud_documento.get(db, doc_id)
    if not doc:
        raise NotFoundError("Documento no encontrado.")
    _assert_acceso(current_user, str(doc.cliente_id))
    if not doc.acuse_recibido_url or not os.path.exists(doc.acuse_recibido_url):
        raise NotFoundError("No hay acuse de recibido registrado.")
    return FileResponse(doc.acuse_recibido_url, filename=doc.acuse_recibido_nombre or "acuse.pdf")
