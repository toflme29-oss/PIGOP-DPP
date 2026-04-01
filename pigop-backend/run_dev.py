"""
Script de arranque para desarrollo local SIN Docker.
Usa SQLite en lugar de PostgreSQL.

Fase 2: crea también validaciones_depp y reglas_normativas,
y siembra el catálogo inicial de reglas normativas.
"""
import sys, os, uuid

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ── Cargar .env primero (preservar valores reales, ej: GEMINI_API_KEY) ────────
_env_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
if os.path.exists(_env_file):
    with open(_env_file) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _v = _line.split("=", 1)
                os.environ.setdefault(_k.strip(), _v.strip())

# ── Variables de entorno ANTES de cualquier import de la app ─────────────────
# DATABASE_URL se fuerza a SQLite para desarrollo local
os.environ["DATABASE_URL"] = "sqlite+aiosqlite:///./pigop_dev.db"
os.environ.setdefault("SECRET_KEY", "dev-secret-key-pigop-2026")
os.environ.setdefault("DEBUG", "True")
os.environ.setdefault("GCS_BUCKET", "pigop-documents-dev")
os.environ.setdefault("GCS_PROJECT_ID", "demo-project")
os.environ.setdefault("GEMINI_API_KEY", "placeholder")   # solo si no está en .env
os.environ.setdefault("SAP_MOCK_MODE", "True")
os.environ.setdefault("STORAGE_BACKEND", "local")
os.environ.setdefault("SUPERADMIN_EMAIL", "admin@pigop.gob.mx")
os.environ.setdefault("SUPERADMIN_PASSWORD", "Admin.2026!")
os.environ["ALLOWED_ORIGINS"] = (
    '["http://localhost:3000","http://localhost:5173","http://localhost:5174",'
    '"http://localhost:8000","http://192.168.10.232:5173",'
    '"http://192.168.10.232:5174","http://192.168.10.232:8000"]'
)

# ── Imports post-config ───────────────────────────────────────────────────────
from sqlalchemy import create_engine, text
import app.models          # noqa — registra todos los modelos con Base.metadata
from app.core.database import Base
from app.core.security import get_password_hash

DB_SYNC = "sqlite:///./pigop_dev.db"
sync_engine = create_engine(DB_SYNC, echo=False)

# ── Crear / actualizar tablas ─────────────────────────────────────────────────
print("🔧 Creando/actualizando tablas SQLite...")
Base.metadata.create_all(sync_engine)
print("✅ Tablas listas")

# ── Migración de documentos_oficiales (agrega columnas nuevas si no existen) ──
_NUEVAS_COLUMNAS_DOC = [
    "ALTER TABLE documentos_oficiales ADD COLUMN flujo VARCHAR(20) DEFAULT 'emitido'",
    "ALTER TABLE documentos_oficiales ADD COLUMN numero_oficio_origen VARCHAR(150)",
    "ALTER TABLE documentos_oficiales ADD COLUMN remitente_nombre VARCHAR(200)",
    "ALTER TABLE documentos_oficiales ADD COLUMN remitente_cargo VARCHAR(200)",
    "ALTER TABLE documentos_oficiales ADD COLUMN remitente_dependencia VARCHAR(200)",
    "ALTER TABLE documentos_oficiales ADD COLUMN fecha_recibido VARCHAR(10)",
    "ALTER TABLE documentos_oficiales ADD COLUMN fecha_limite VARCHAR(10)",
    "ALTER TABLE documentos_oficiales ADD COLUMN prioridad VARCHAR(20) DEFAULT 'normal'",
    "ALTER TABLE documentos_oficiales ADD COLUMN ocr_procesado INTEGER DEFAULT 0",
    "ALTER TABLE documentos_oficiales ADD COLUMN texto_extraido_ocr TEXT",
    "ALTER TABLE documentos_oficiales ADD COLUMN datos_extraidos_ia TEXT",
    "ALTER TABLE documentos_oficiales ADD COLUMN sugerencia_area_codigo VARCHAR(10)",
    "ALTER TABLE documentos_oficiales ADD COLUMN sugerencia_area_nombre VARCHAR(200)",
    "ALTER TABLE documentos_oficiales ADD COLUMN sugerencia_fundamento TEXT",
    "ALTER TABLE documentos_oficiales ADD COLUMN sugerencia_plazo_dias INTEGER",
    "ALTER TABLE documentos_oficiales ADD COLUMN confianza_clasificacion REAL",
    "ALTER TABLE documentos_oficiales ADD COLUMN regla_turno_codigo VARCHAR(30)",
    "ALTER TABLE documentos_oficiales ADD COLUMN genera_tramite VARCHAR(50)",
    "ALTER TABLE documentos_oficiales ADD COLUMN area_turno VARCHAR(10)",
    "ALTER TABLE documentos_oficiales ADD COLUMN area_turno_nombre VARCHAR(200)",
    "ALTER TABLE documentos_oficiales ADD COLUMN area_turno_confirmada INTEGER DEFAULT 0",
    "ALTER TABLE documentos_oficiales ADD COLUMN turnado_por_id VARCHAR(36)",
    "ALTER TABLE documentos_oficiales ADD COLUMN turnado_en DATETIME",
    "ALTER TABLE documentos_oficiales ADD COLUMN borrador_respuesta TEXT",
    "ALTER TABLE documentos_oficiales ADD COLUMN folio_respuesta VARCHAR(100)",
    "ALTER TABLE documentos_oficiales ADD COLUMN fecha_respuesta VARCHAR(50)",
    "ALTER TABLE documentos_oficiales ADD COLUMN referencia_elaboro VARCHAR(50)",
    "ALTER TABLE documentos_oficiales ADD COLUMN referencia_reviso VARCHAR(50)",
    "ALTER TABLE documentos_oficiales ADD COLUMN tabla_imagen_url VARCHAR(500)",
    "ALTER TABLE documentos_oficiales ADD COLUMN tabla_imagen_nombre VARCHAR(200)",
    "ALTER TABLE documentos_oficiales ADD COLUMN tabla_datos_json JSON",
    "ALTER TABLE documentos_oficiales ADD COLUMN certificacion_id VARCHAR(36)",
    "ALTER TABLE documentos_oficiales ADD COLUMN modulo_externo_estado VARCHAR(50)",
    "ALTER TABLE documentos_oficiales ADD COLUMN modulo_externo_ref VARCHAR(200)",
    "ALTER TABLE documentos_oficiales ADD COLUMN upp_solicitante VARCHAR(20)",
    "ALTER TABLE documentos_oficiales ADD COLUMN termino_contestacion VARCHAR(150)",
    "ALTER TABLE documentos_oficiales ADD COLUMN subdirector_nombre VARCHAR(200)",
    "ALTER TABLE documentos_oficiales ADD COLUMN jefe_departamento_nombre VARCHAR(200)",
    "ALTER TABLE documentos_oficiales ADD COLUMN firmado_digitalmente INTEGER DEFAULT 0",
    "ALTER TABLE documentos_oficiales ADD COLUMN firma_metadata TEXT",
    # ── Instrucciones de turno ──
    "ALTER TABLE documentos_oficiales ADD COLUMN instrucciones_turno TEXT",
    # ── Para conocimiento ──
    "ALTER TABLE documentos_oficiales ADD COLUMN requiere_respuesta INTEGER DEFAULT 1",
    # ── Devolución y versionado (Fase firma por lote) ──
    "ALTER TABLE documentos_oficiales ADD COLUMN version INTEGER DEFAULT 1",
    "ALTER TABLE documentos_oficiales ADD COLUMN devuelto_por_id VARCHAR(36)",
    "ALTER TABLE documentos_oficiales ADD COLUMN devuelto_en DATETIME",
    "ALTER TABLE documentos_oficiales ADD COLUMN motivo_devolucion TEXT",
    # ── Despacho por secretaria ──
    "ALTER TABLE documentos_oficiales ADD COLUMN despachado INTEGER DEFAULT 0",
    "ALTER TABLE documentos_oficiales ADD COLUMN despachado_por_id VARCHAR(36)",
    "ALTER TABLE documentos_oficiales ADD COLUMN despachado_en DATETIME",
    # ── Acuse de recibido ──
    "ALTER TABLE documentos_oficiales ADD COLUMN acuse_recibido_url VARCHAR(500)",
    "ALTER TABLE documentos_oficiales ADD COLUMN acuse_recibido_nombre VARCHAR(200)",
    "ALTER TABLE documentos_oficiales ADD COLUMN acuse_recibido_fecha VARCHAR(100)",
    "ALTER TABLE documentos_oficiales ADD COLUMN acuse_registrado_en DATETIME",
    "ALTER TABLE documentos_oficiales ADD COLUMN acuse_registrado_por_id VARCHAR(36)",
]
with sync_engine.connect() as _mc:
    for _stmt in _NUEVAS_COLUMNAS_DOC:
        try:
            _mc.execute(text(_stmt))
        except Exception:
            pass  # la columna ya existe
    _mc.commit()
print("✅ Columnas documentos_oficiales actualizadas")

# ── Catálogo de funcionarios y UPPs ──────────────────────────────────────────
with sync_engine.connect() as _catc:
    _catc.execute(text('''CREATE TABLE IF NOT EXISTS catalogo_funcionarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_upp VARCHAR(10) NOT NULL,
        nombre_upp VARCHAR(255) NOT NULL,
        codigo_ur VARCHAR(10),
        nombre_ur VARCHAR(255),
        nombre_titular VARCHAR(255),
        cargo VARCHAR(255),
        estatus VARCHAR(50) DEFAULT 'VIGENTE'
    )'''))
    _catc.commit()
    # Importar datos si la tabla está vacía
    _count = _catc.execute(text("SELECT COUNT(*) FROM catalogo_funcionarios")).fetchone()[0]
    if _count == 0:
        _cat_path = os.path.join(os.path.dirname(__file__), "..", "OFICIOS PRUEBA", "CATALOGO FUNCIONARIOS UPPS", "CATALOGO FUNCIONARIOS UPPS.xlsx")
        # Intentar ruta alternativa
        _cat_paths = [
            _cat_path,
            "/Users/mafm/Documents/Documents/SECRETARÍA DE FINANZAS MICH/IA DPP/OFICIOS PRUEBA/CATALOGO FUNCIONARIOS UPPS/CATALOGO FUNCIONARIOS UPPS.xlsx",
        ]
        _cat_file = None
        for _cp in _cat_paths:
            if os.path.exists(_cp):
                _cat_file = _cp
                break
        if _cat_file:
            try:
                import openpyxl as _opx
                _wb = _opx.load_workbook(_cat_file, data_only=True)
                _ws = _wb['2021-2027']
                _ins = 0
                for _row in _ws.iter_rows(min_row=2, max_row=_ws.max_row, values_only=True):
                    _id_upp = str(_row[0]).strip() if _row[0] else ''
                    _desc_upp = str(_row[1]).strip() if _row[1] else ''
                    _cod_ur = str(_row[2]).strip() if _row[2] else ''
                    _desc_ur = str(_row[3]).strip() if _row[3] else ''
                    _titular = str(_row[4]).strip() if _row[4] else ''
                    _est = str(_row[5]).strip() if _row[5] else ''
                    if not _id_upp or _id_upp == 'None' or _est != 'VIGENTE':
                        continue
                    _catc.execute(text(
                        "INSERT INTO catalogo_funcionarios (codigo_upp,nombre_upp,codigo_ur,nombre_ur,nombre_titular,cargo,estatus) "
                        "VALUES (:u,:nu,:ur,:nur,:tit,:car,:est)"
                    ), {"u": _id_upp, "nu": _desc_upp, "ur": _cod_ur, "nur": _desc_ur, "tit": _titular, "car": _desc_ur, "est": _est})
                    _ins += 1
                _catc.commit()
                print(f"✅ Catálogo: {_ins} funcionarios importados")
            except Exception as _e:
                print(f"⚠️  Error importando catálogo: {_e}")
        else:
            print("⚠️  Archivo catálogo no encontrado — tabla vacía")
    else:
        print(f"✅ Catálogo: {_count} funcionarios existentes")

# ── Migración de usuarios (columnas nuevas) ─────────────────────────────────
_NUEVAS_COLUMNAS_USR = [
    "ALTER TABLE usuarios ADD COLUMN modulos_acceso TEXT DEFAULT '[]'",
]
with sync_engine.connect() as _uc:
    for _stmt in _NUEVAS_COLUMNAS_USR:
        try:
            _uc.execute(text(_stmt))
        except Exception:
            pass  # la columna ya existe
    _uc.commit()
print("✅ Columnas usuarios actualizadas")

# ── Crear tablas de bóveda de certificados y bitácora de firma ──────────────
_TABLAS_BOVEDA = [
    """CREATE TABLE IF NOT EXISTS certificados_firma (
        id VARCHAR(36) PRIMARY KEY,
        usuario_id VARCHAR(36) NOT NULL UNIQUE,
        cer_data TEXT NOT NULL,
        key_data_cifrada TEXT NOT NULL,
        key_iv VARCHAR(64) NOT NULL,
        key_tag VARCHAR(64) NOT NULL,
        rfc VARCHAR(20) NOT NULL,
        nombre_titular VARCHAR(300) NOT NULL,
        numero_serie VARCHAR(100) NOT NULL,
        valido_desde DATETIME,
        valido_hasta DATETIME,
        emisor VARCHAR(300),
        activo INTEGER DEFAULT 1,
        registrado_en DATETIME DEFAULT CURRENT_TIMESTAMP,
        ultima_firma_en DATETIME,
        total_firmas INTEGER DEFAULT 0,
        FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
    )""",
    """CREATE TABLE IF NOT EXISTS bitacora_firma (
        id VARCHAR(36) PRIMARY KEY,
        usuario_id VARCHAR(36) NOT NULL,
        accion VARCHAR(50) NOT NULL,
        documento_id VARCHAR(36),
        lote_firma_id VARCHAR(36),
        rfc_certificado VARCHAR(20),
        numero_serie VARCHAR(100),
        hash_documento VARCHAR(128),
        ip_origen VARCHAR(45),
        exitoso INTEGER DEFAULT 1,
        detalle TEXT,
        timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
    )""",
]
with sync_engine.connect() as _bc:
    for _sql in _TABLAS_BOVEDA:
        try:
            _bc.execute(text(_sql))
        except Exception:
            pass
    _bc.commit()
print("✅ Tablas bóveda certificados y bitácora firma listas")

EMAIL = os.environ["SUPERADMIN_EMAIL"]
PWD   = os.environ["SUPERADMIN_PASSWORD"]

# ── Datos semilla ─────────────────────────────────────────────────────────────
with sync_engine.connect() as conn:

    # ── Cliente DPP ───────────────────────────────────────────────────────────
    if not conn.execute(
        text("SELECT 1 FROM clientes WHERE codigo_upp='DPP'")
    ).fetchone():
        conn.execute(text(
            "INSERT INTO clientes "
            "(id, codigo_upp, nombre, tipo, activo, configuracion) "
            "VALUES (:id, 'DPP', 'Dirección de Programación y Presupuesto', "
            "        'centralizada', 1, '{}')"
        ), {"id": str(uuid.uuid4())})
        conn.commit()
        print("✅ Cliente DPP creado")

    # ── Superadmin ────────────────────────────────────────────────────────────
    if not conn.execute(
        text("SELECT 1 FROM usuarios WHERE email=:e"), {"e": EMAIL}
    ).fetchone():
        conn.execute(text(
            "INSERT INTO usuarios "
            "(id, email, password_hash, nombre_completo, rol, activo, modulos_acceso) "
            "VALUES (:id, :email, :pwd, 'Administrador PIGOP', 'superadmin', 1, '[]')"
        ), {"id": str(uuid.uuid4()), "email": EMAIL, "pwd": get_password_hash(PWD)})
        conn.commit()
        print(f"✅ Superadmin: {EMAIL}")

    # ── Obtener cliente_id DPP ───────────────────────────────────────────────
    _dpp_row = conn.execute(text("SELECT id FROM clientes WHERE codigo_upp='DPP'")).fetchone()
    _dpp_id = _dpp_row[0] if _dpp_row else None

    # ── Usuarios de prueba (tabla de permisos Gestión Documental v4) ────────
    _usuarios_gd = [
        # (email, password, nombre, rol)
        ("director@pigop.gob.mx", "Dir2026!", "Mtro. Marco Antonio Flores Mejía", "admin_cliente"),
        ("secretaria@pigop.gob.mx", "Sec2026!", "Berenice Huerta Silva", "secretaria"),
        ("asesor@pigop.gob.mx", "Ase2026!", "René Emilio Rico García", "asesor"),
        ("subcep@pigop.gob.mx", "Sub2026!", "Eduardo Cortés Jaramillo", "subdirector"),
        ("jdcpres@pigop.gob.mx", "Jef2026!", "Blanca Esthela Ortíz Soto", "jefe_depto"),
        ("jdregej@pigop.gob.mx", "Jef2026!", "Luis Alberto Sánchez León", "jefe_depto"),
        ("subpyf@pigop.gob.mx", "Sub2026!", "José Luis Pardo Escutia", "subdirector"),
        ("jdasyp@pigop.gob.mx", "Jef2026!", "Seomara Mendoza Cárdenas", "jefe_depto"),
        ("jdfyn@pigop.gob.mx", "Jef2026!", "Hugo Díaz Arechiaga", "jefe_depto"),
        ("auditor@pigop.gob.mx", "Aud2026!", "Auditor SGC", "auditor"),
    ]
    for _em, _pw, _nm, _rl in _usuarios_gd:
        existing = conn.execute(text("SELECT 1 FROM usuarios WHERE email=:e"), {"e": _em}).fetchone()
        if not existing:
            conn.execute(text(
                "INSERT INTO usuarios "
                "(id, email, password_hash, nombre_completo, rol, activo, cliente_id, modulos_acceso) "
                "VALUES (:id, :email, :pwd, :nombre, :rol, 1, :cid, :modulos)"
            ), {"id": str(uuid.uuid4()), "email": _em, "pwd": get_password_hash(_pw),
                "nombre": _nm, "rol": _rl, "cid": _dpp_id, "modulos": '["todos"]'})
        else:
            conn.execute(text(
                "UPDATE usuarios SET password_hash=:pwd, nombre_completo=:nombre, rol=:rol, cliente_id=:cid WHERE email=:e"
            ), {"pwd": get_password_hash(_pw), "nombre": _nm, "rol": _rl, "cid": _dpp_id, "e": _em})
    conn.commit()
    print(f"✅ {len(_usuarios_gd)} usuarios de Gestión Documental creados/actualizados")

    # ── Reglas normativas iniciales ───────────────────────────────────────────
    # Campos: codigo, articulo, titulo, descripcion, tipo_validacion,
    #         aplica_clasificacion (JSON str), aplica_capitulo (JSON str),
    #         gravedad, bloquea_aprobacion (0/1), activa (1), version (int)
    import json as _json

    REGLAS = [
        # ---- Estructura ----
        {
            "codigo": "EST-001",
            "articulo": "Art. 25",
            "titulo": "Campos obligatorios del DEPP",
            "descripcion": "El DEPP debe contener folio, UPP, ejercicio y monto total.",
            "tipo_validacion": "estructura",
            "aplica_clasificacion": None,
            "aplica_capitulo": None,
            "gravedad": "critico",
            "bloquea_aprobacion": 1,
        },
        {
            "codigo": "EST-002",
            "articulo": "Art. 25",
            "titulo": "Monto total válido",
            "descripcion": "El monto total del DEPP debe ser mayor a cero.",
            "tipo_validacion": "estructura",
            "aplica_clasificacion": None,
            "aplica_capitulo": None,
            "gravedad": "critico",
            "bloquea_aprobacion": 1,
        },
        # ---- Documentos ----
        {
            "codigo": "DOC-I1",
            "articulo": "Art. 39 fracc. I",
            "titulo": "Documentos requeridos clasificación I.1",
            "descripcion": "Clasificación I.1: requiere DEPP + CFDI + Contrato (CTT) + Manifiesto (MCL).",
            "tipo_validacion": "documental",
            "aplica_clasificacion": _json.dumps(["I.1"]),
            "aplica_capitulo": _json.dumps([2000, 3000, 5000, 6000]),
            "gravedad": "critico",
            "bloquea_aprobacion": 1,
        },
        {
            "codigo": "DOC-II1",
            "articulo": "Art. 39 fracc. II.1",
            "titulo": "Documentos requeridos clasificación II.1",
            "descripcion": "Clasificación II.1: requiere DEPP + Acuerdo Único de Reasignación (AUR).",
            "tipo_validacion": "documental",
            "aplica_clasificacion": _json.dumps(["II.1"]),
            "aplica_capitulo": None,
            "gravedad": "critico",
            "bloquea_aprobacion": 1,
        },
        {
            "codigo": "DOC-II2",
            "articulo": "Art. 39 fracc. II.2",
            "titulo": "Documentos requeridos clasificación II.2",
            "descripcion": "Clasificación II.2: requiere DEPP + Formato Único de Comisión (FUC). Solo cap. 3000.",
            "tipo_validacion": "documental",
            "aplica_clasificacion": _json.dumps(["II.2"]),
            "aplica_capitulo": _json.dumps([3000]),
            "gravedad": "critico",
            "bloquea_aprobacion": 1,
        },
        {
            "codigo": "DOC-II3",
            "articulo": "Art. 39 fracc. II.3",
            "titulo": "Documentos requeridos clasificación II.3",
            "descripcion": "Clasificación II.3: requiere DEPP + Póliza Cheque/Transferencia (PCH).",
            "tipo_validacion": "documental",
            "aplica_clasificacion": _json.dumps(["II.3"]),
            "aplica_capitulo": None,
            "gravedad": "critico",
            "bloquea_aprobacion": 1,
        },
        {
            "codigo": "DOC-II4",
            "articulo": "Art. 39 fracc. II.4",
            "titulo": "Documentos requeridos clasificación II.4",
            "descripcion": "Clasificación II.4: requiere DEPP + CFDI + Manifiesto (MCL), sin contrato.",
            "tipo_validacion": "documental",
            "aplica_clasificacion": _json.dumps(["II.4"]),
            "aplica_capitulo": _json.dumps([2000, 3000, 5000]),
            "gravedad": "critico",
            "bloquea_aprobacion": 1,
        },
        # ---- Coherencia ----
        {
            "codigo": "COH-001",
            "articulo": "Art. 40",
            "titulo": "Compatibilidad capítulo-clasificación",
            "descripcion": "El capítulo presupuestal debe ser compatible con la clasificación del DEPP.",
            "tipo_validacion": "presupuestal",
            "aplica_clasificacion": None,
            "aplica_capitulo": None,
            "gravedad": "alto",
            "bloquea_aprobacion": 0,
        },
        {
            "codigo": "COH-002",
            "articulo": "Art. 40",
            "titulo": "Monto mayor a $500,000 — verificar licitación",
            "descripcion": (
                "Montos superiores a $500,000 deben contar con proceso de licitación "
                "o justificación de excepción (Art. 41 LAASSP)."
            ),
            "tipo_validacion": "presupuestal",
            "aplica_clasificacion": None,
            "aplica_capitulo": None,
            "gravedad": "medio",
            "bloquea_aprobacion": 0,
        },
        {
            "codigo": "COH-003",
            "articulo": "Art. 40",
            "titulo": "CFDI en capítulo 1000 sin contrato",
            "descripcion": (
                "Capítulo 1000 (Servicios Personales) con CFDI externo sin contrato "
                "requiere verificación adicional."
            ),
            "tipo_validacion": "presupuestal",
            "aplica_clasificacion": None,
            "aplica_capitulo": _json.dumps([1000]),
            "gravedad": "medio",
            "bloquea_aprobacion": 0,
        },
        # ---- Clasificación ----
        {
            "codigo": "CLA-001",
            "articulo": "Art. 39",
            "titulo": "Clasificación normativa determinable",
            "descripcion": (
                "El sistema debe poder determinar la clasificación normativa del DEPP "
                "a partir de los documentos adjuntos."
            ),
            "tipo_validacion": "documental",
            "aplica_clasificacion": None,
            "aplica_capitulo": None,
            "gravedad": "alto",
            "bloquea_aprobacion": 0,
        },
        {
            "codigo": "CLA-002",
            "articulo": "Art. 39",
            "titulo": "Coincidencia de clasificación capturada",
            "descripcion": (
                "La clasificación capturada manualmente debe coincidir con "
                "la detectada automáticamente por el sistema."
            ),
            "tipo_validacion": "documental",
            "aplica_clasificacion": None,
            "aplica_capitulo": None,
            "gravedad": "medio",
            "bloquea_aprobacion": 0,
        },
    ]

    for regla in REGLAS:
        existe = conn.execute(
            text("SELECT 1 FROM reglas_normativas WHERE codigo=:c"),
            {"c": regla["codigo"]},
        ).fetchone()
        if not existe:
            conn.execute(
                text(
                    "INSERT INTO reglas_normativas "
                    "(id, codigo, articulo, titulo, descripcion, tipo_validacion, "
                    " aplica_clasificacion, aplica_capitulo, gravedad, "
                    " bloquea_aprobacion, activa, version) "
                    "VALUES (:id, :codigo, :articulo, :titulo, :descripcion, :tipo_validacion, "
                    "        :aplica_clasificacion, :aplica_capitulo, :gravedad, "
                    "        :bloquea_aprobacion, 1, 2)"
                ),
                {
                    "id": str(uuid.uuid4()),
                    "codigo": regla["codigo"],
                    "articulo": regla["articulo"],
                    "titulo": regla["titulo"],
                    "descripcion": regla["descripcion"],
                    "tipo_validacion": regla["tipo_validacion"],
                    "aplica_clasificacion": regla["aplica_clasificacion"],
                    "aplica_capitulo": regla["aplica_capitulo"],
                    "gravedad": regla["gravedad"],
                    "bloquea_aprobacion": regla["bloquea_aprobacion"],
                },
            )
    conn.commit()
    total_reglas = conn.execute(
        text("SELECT COUNT(*) FROM reglas_normativas")
    ).fetchone()[0]
    print(f"✅ Reglas normativas: {total_reglas} registros")

sync_engine.dispose()

# ── Banner de inicio ──────────────────────────────────────────────────────────
print()
print("=" * 60)
print("🚀  PIGOP Backend — Fase 3 (Gemini AI + OCR + Validación IA)")
print("=" * 60)
print(f"   Local:      http://localhost:8000")
print(f"   Red:        http://192.168.10.232:8000")
print(f"   Swagger UI: http://localhost:8000/docs")
print(f"   Health:     http://localhost:8000/health")
print()
print(f"   Email:      {EMAIL}")
print(f"   Password:   {PWD}")
print("=" * 60)
print()

import uvicorn
uvicorn.run("app.main:app", host="0.0.0.0", port=8000, reload=False)
